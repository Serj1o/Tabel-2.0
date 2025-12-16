import asyncio
import logging
import os
import math
import tempfile
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from calendar import monthrange
from enum import Enum
from typing import Optional, List, Tuple
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
logger.error("🚨 BOT REACHED START_POLLING")

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from aiogram import Bot, Dispatcher, Router, F
from aiogram.filters import Command
from aiogram.types import (
    Message,
    CallbackQuery,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    FSInputFile,
)
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext

from apscheduler.schedulers.asyncio import AsyncIOScheduler
import aiosmtplib
from email.message import EmailMessage

# =========================
# ENV / CONFIG
# =========================

load_dotenv()

API_TOKEN = os.getenv("API_TOKEN")
DATABASE_URL = os.getenv("DATABASE_URL")

TZ = ZoneInfo("Europe/Amsterdam")
GEO_REQUIRED = os.getenv("GEO_REQUIRED", "false").lower() == "true"

ADMIN_IDS_ENV = os.getenv("ADMIN_IDS", "467500951")
ADMIN_IDS = [int(x.strip()) for x in ADMIN_IDS_ENV.split(",") if x.strip().isdigit()]

SMTP_HOST = os.getenv("SMTP_HOST")
SMTP_PORT = int(os.getenv("SMTP_PORT", "465"))
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASS = os.getenv("SMTP_PASS")
REPORT_EMAILS = [x.strip() for x in (os.getenv("REPORT_EMAILS", "")).split(",") if x.strip()]

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger("timesheet-bot")

# =========================
# ROLES / STATUSES
# =========================

class Role(str, Enum):
    ADMIN = "admin"
    EMPLOYEE = "employee"

class DayStatus(str, Enum):
    WORK = "work"
    SICK = "sick"

# =========================
# DB
# =========================

pool: Optional[asyncpg.Pool] = None

async def db_init():
    global pool
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL not set")

    pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=10)

    async with pool.acquire() as con:
        await con.execute("""
        CREATE TABLE IF NOT EXISTS users (
            tg_id BIGINT PRIMARY KEY,
            fio TEXT,
            position TEXT,
            phone TEXT,
            role TEXT NOT NULL DEFAULT 'employee',
            approved BOOLEAN NOT NULL DEFAULT FALSE,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        );
        """)

        await con.execute("""
        CREATE TABLE IF NOT EXISTS objects (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            lat DOUBLE PRECISION NOT NULL,
            lon DOUBLE PRECISION NOT NULL,
            radius_m INTEGER NOT NULL DEFAULT 200
        );
        """)

        await con.execute("""
        CREATE TABLE IF NOT EXISTS shifts (
            id SERIAL PRIMARY KEY,
            tg_id BIGINT NOT NULL REFERENCES users(tg_id) ON DELETE CASCADE,
            day DATE NOT NULL,
            object_id INTEGER REFERENCES objects(id) ON DELETE SET NULL,
            start_ts TIMESTAMPTZ,
            end_ts TIMESTAMPTZ,
            start_lat DOUBLE PRECISION,
            start_lon DOUBLE PRECISION,
            end_lat DOUBLE PRECISION,
            end_lon DOUBLE PRECISION,
            hours NUMERIC(4,2),
            status TEXT NOT NULL DEFAULT 'work',
            UNIQUE(tg_id, day)
        );
        """)

async def db_user_get(tg_id: int) -> Optional[asyncpg.Record]:
    async with pool.acquire() as con:
        return await con.fetchrow("SELECT * FROM users WHERE tg_id=$1", tg_id)

async def db_user_ensure(tg_id: int) -> asyncpg.Record:
    async with pool.acquire() as con:
        await con.execute("""
            INSERT INTO users (tg_id, role, approved)
            VALUES ($1, $2, $3)
            ON CONFLICT (tg_id) DO NOTHING
        """, tg_id, Role.ADMIN if tg_id in ADMIN_IDS else Role.EMPLOYEE, True if tg_id in ADMIN_IDS else False)
        return await con.fetchrow("SELECT * FROM users WHERE tg_id=$1", tg_id)

async def db_user_update(tg_id: int, fio: Optional[str]=None, position: Optional[str]=None, phone: Optional[str]=None):
    async with pool.acquire() as con:
        await con.execute("""
            UPDATE users
            SET fio = COALESCE($2, fio),
                position = COALESCE($3, position),
                phone = COALESCE($4, phone)
            WHERE tg_id=$1
        """, tg_id, fio, position, phone)

async def db_user_set_approved(tg_id: int, approved: bool):
    async with pool.acquire() as con:
        await con.execute("UPDATE users SET approved=$2 WHERE tg_id=$1", tg_id, approved)

async def db_users_list_approved() -> List[asyncpg.Record]:
    async with pool.acquire() as con:
        return await con.fetch("SELECT * FROM users WHERE approved=TRUE ORDER BY fio NULLS LAST")

async def db_objects_list() -> List[asyncpg.Record]:
    async with pool.acquire() as con:
        return await con.fetch("SELECT * FROM objects ORDER BY name")

async def db_object_add(name: str, lat: float, lon: float, radius_m: int):
    async with pool.acquire() as con:
        await con.execute(
            "INSERT INTO objects(name, lat, lon, radius_m) VALUES ($1,$2,$3,$4) ON CONFLICT(name) DO NOTHING",
            name, lat, lon, radius_m
        )

async def db_object_get(object_id: int) -> Optional[asyncpg.Record]:
    async with pool.acquire() as con:
        return await con.fetchrow("SELECT * FROM objects WHERE id=$1", object_id)

async def db_shift_get(tg_id: int, day: date) -> Optional[asyncpg.Record]:
    async with pool.acquire() as con:
        return await con.fetchrow("SELECT * FROM shifts WHERE tg_id=$1 AND day=$2", tg_id, day)

async def db_shift_ensure(tg_id: int, day: date):
    async with pool.acquire() as con:
        await con.execute("""
            INSERT INTO shifts (tg_id, day)
            VALUES ($1, $2)
            ON CONFLICT (tg_id, day) DO NOTHING
        """, tg_id, day)

async def db_shift_set_object(tg_id: int, day: date, object_id: int):
    await db_shift_ensure(tg_id, day)
    async with pool.acquire() as con:
        await con.execute("UPDATE shifts SET object_id=$3 WHERE tg_id=$1 AND day=$2", tg_id, day, object_id)

async def db_shift_set_sick(tg_id: int, day: date, sick: bool):
    await db_shift_ensure(tg_id, day)
    async with pool.acquire() as con:
        if sick:
            await con.execute("""
                UPDATE shifts
                SET status='sick', hours=NULL, start_ts=NULL, end_ts=NULL
                WHERE tg_id=$1 AND day=$2
            """, tg_id, day)
        else:
            await con.execute("""
                UPDATE shifts
                SET status='work'
                WHERE tg_id=$1 AND day=$2
            """, tg_id, day)

async def db_shift_start(tg_id: int, day: date, ts: datetime, lat: Optional[float], lon: Optional[float]):
    await db_shift_ensure(tg_id, day)
    async with pool.acquire() as con:
        await con.execute("""
            UPDATE shifts
            SET status='work',
                start_ts=$3, end_ts=NULL, hours=NULL,
                start_lat=$4, start_lon=$5
            WHERE tg_id=$1 AND day=$2
        """, tg_id, day, ts, lat, lon)

async def db_shift_end(tg_id: int, day: date, ts: datetime, lat: Optional[float], lon: Optional[float], hours: float):
    async with pool.acquire() as con:
        await con.execute("""
            UPDATE shifts
            SET end_ts=$3, end_lat=$4, end_lon=$5, hours=$6
            WHERE tg_id=$1 AND day=$2
        """, tg_id, day, ts, lat, lon, hours)

async def db_month_shifts(year: int, month: int) -> List[asyncpg.Record]:
    d1 = date(year, month, 1)
    d2 = date(year, month, monthrange(year, month)[1])
    async with pool.acquire() as con:
        return await con.fetch("""
            SELECT s.*, u.fio, u.position
            FROM shifts s
            JOIN users u ON u.tg_id = s.tg_id
            WHERE s.day BETWEEN $1 AND $2 AND u.approved=TRUE
            ORDER BY u.fio NULLS LAST, s.day
        """, d1, d2)

# =========================
# HELPERS
# =========================

def now() -> datetime:
    return datetime.now(TZ)

def today() -> date:
    return now().date()

def haversine_m(lat1, lon1, lat2, lon2) -> float:
    # meters
    R = 6371000.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dl/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def round_up_half_hour(hours: float) -> float:
    # округление вверх до 0.5 часа
    return math.ceil(hours * 2) / 2.0

def calc_hours(start_ts: datetime, end_ts: datetime) -> float:
    h = (end_ts - start_ts).total_seconds() / 3600.0
    h = max(0.0, h)
    h = round_up_half_hour(h)
    return min(8.0, h)

def require_approved(user: asyncpg.Record) -> bool:
    return bool(user["approved"])

# =========================
# KEYBOARDS
# =========================

def main_kb(role: str):
    kb = [
        ["🏗 Выбрать объект"],
        ["✅ Пришел", "⛔ Ушел"],
        ["🤒 Болел", "ℹ️ Мой статус"],
    ]
    if role == Role.ADMIN:
        kb.append(["📋 Сотрудники", "🏢 Объекты", "📤 Сформировать табель"])
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)

kb_geo = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="📍 Отправить геолокацию", request_location=True)]],
    resize_keyboard=True,
)

# =========================
# FSM
# =========================

class AuthFSM(StatesGroup):
    fio = State()
    position = State()

class ObjectFSM(StatesGroup):
    name = State()
    lat = State()
    lon = State()
    radius = State()

# =========================
# ROUTER
# =========================

router = Router()

# =========================
# START / AUTH
# =========================

@router.message(Command("start"))
async def cmd_start(msg: Message, state: FSMContext):
    u = await db_user_ensure(msg.from_user.id)
    await state.clear()

    if require_approved(u):
        await msg.answer("🏠 Главное меню", reply_markup=main_kb(u["role"]))
        return

    await msg.answer(
        "👋 Добро пожаловать!\n\n"
        "Для доступа нужна регистрация и одобрение админа.\n"
        "Введи ФИО (например: Иванов Иван):"
    )
    await state.set_state(AuthFSM.fio)

@router.message(AuthFSM.fio)
async def auth_fio(msg: Message, state: FSMContext):
    fio = (msg.text or "").strip()
    if len(fio) < 3:
        await msg.answer("ФИО слишком короткое. Введи ещё раз:")
        return
    await db_user_update(msg.from_user.id, fio=fio)
    await msg.answer("Введи должность (например: Монтажник):")
    await state.set_state(AuthFSM.position)

@router.message(AuthFSM.position)
async def auth_position(msg: Message, state: FSMContext):
    pos = (msg.text or "").strip()
    await db_user_update(msg.from_user.id, position=pos)
    await state.clear()

    u = await db_user_get(msg.from_user.id)
    fio = u["fio"] or "—"
    position = u["position"] or "—"

    # уведомляем админов
    for admin_id in ADMIN_IDS:
        try:
            await msg.bot.send_message(
                admin_id,
                f"🔔 Новая заявка\n\nФИО: {fio}\nДолжность: {position}\nTG: {msg.from_user.id}",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[[
                        InlineKeyboardButton(text="✅ Одобрить", callback_data=f"approve:{msg.from_user.id}"),
                        InlineKeyboardButton(text="❌ Отклонить", callback_data=f"reject:{msg.from_user.id}"),
                    ]]
                )
            )
        except Exception:
            logger.exception("Failed to notify admin %s", admin_id)

    await msg.answer("⏳ Заявка отправлена администратору. Ожидай подтверждения.")

@router.callback_query(F.data.startswith("approve:"))
async def cb_approve(cb: CallbackQuery):
    if cb.from_user.id not in ADMIN_IDS:
        await cb.answer("Нет прав", show_alert=True)
        return

    tg_id = int(cb.data.split(":")[1])
    await db_user_set_approved(tg_id, True)
    await cb.answer("Одобрено ✅")

    try:
        u = await db_user_get(tg_id)
        await cb.bot.send_message(tg_id, "✅ Доступ одобрен. Открой /start")
        if u and u["role"] != Role.ADMIN:
            # роль оставляем employee, но approved = True
            pass
    except Exception:
        logger.exception("Failed to message approved user")

@router.callback_query(F.data.startswith("reject:"))
async def cb_reject(cb: CallbackQuery):
    if cb.from_user.id not in ADMIN_IDS:
        await cb.answer("Нет прав", show_alert=True)
        return
    tg_id = int(cb.data.split(":")[1])
    await db_user_set_approved(tg_id, False)
    await cb.answer("Отклонено ❌")
    try:
        await cb.bot.send_message(tg_id, "❌ Доступ не одобрен. Если это ошибка — напиши администратору.")
    except Exception:
        pass

# =========================
# OBJECTS (SELECT + ADMIN CRUD-lite)
# =========================

@router.message(F.text == "🏗 Выбрать объект")
async def pick_object(msg: Message):
    u = await db_user_ensure(msg.from_user.id)
    if not require_approved(u):
        return

    objs = await db_objects_list()
    if not objs:
        await msg.answer("Пока нет объектов. Админ должен добавить хотя бы один.")
        return

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=o["name"], callback_data=f"obj:{o['id']}")]
        for o in objs
    ])
    await msg.answer("Выбери объект на сегодня:", reply_markup=kb)

@router.callback_query(F.data.startswith("obj:"))
async def cb_set_object(cb: CallbackQuery):
    u = await db_user_ensure(cb.from_user.id)
    if not require_approved(u):
        await cb.answer("Нет доступа", show_alert=True)
        return

    object_id = int(cb.data.split(":")[1])
    obj = await db_object_get(object_id)
    if not obj:
        await cb.answer("Объект не найден", show_alert=True)
        return

    await db_shift_set_object(cb.from_user.id, today(), object_id)
    await cb.answer("Объект выбран ✅")
    await cb.message.answer(f"🏗 Объект на сегодня: {obj['name']}", reply_markup=main_kb(u["role"]))

@router.message(F.text == "🏢 Объекты")
async def admin_objects(msg: Message, state: FSMContext):
    u = await db_user_ensure(msg.from_user.id)
    if u["role"] != Role.ADMIN:
        return

    objs = await db_objects_list()
    text = "🏢 Объекты:\n" + ("\n".join([f"- {o['name']} (r={o['radius_m']}м)" for o in objs]) if objs else "— нет —")
    text += "\n\nЧтобы добавить объект — отправь: “Добавить объект”"
    await msg.answer(text, reply_markup=main_kb(u["role"]))

@router.message(F.text.lower() == "добавить объект")
async def admin_add_object_start(msg: Message, state: FSMContext):
    u = await db_user_ensure(msg.from_user.id)
    if u["role"] != Role.ADMIN:
        return
    await state.set_state(ObjectFSM.name)
    await msg.answer("Введи название объекта:")

@router.message(ObjectFSM.name)
async def admin_add_object_name(msg: Message, state: FSMContext):
    await state.update_data(name=(msg.text or "").strip())
    await state.set_state(ObjectFSM.lat)
    await msg.answer("Введи широту (lat), например 52.370216:")

@router.message(ObjectFSM.lat)
async def admin_add_object_lat(msg: Message, state: FSMContext):
    try:
        lat = float((msg.text or "").replace(",", "."))
    except ValueError:
        await msg.answer("Не похоже на число. Введи широту ещё раз:")
        return
    await state.update_data(lat=lat)
    await state.set_state(ObjectFSM.lon)
    await msg.answer("Введи долготу (lon), например 4.895168:")

@router.message(ObjectFSM.lon)
async def admin_add_object_lon(msg: Message, state: FSMContext):
    try:
        lon = float((msg.text or "").replace(",", "."))
    except ValueError:
        await msg.answer("Не похоже на число. Введи долготу ещё раз:")
        return
    await state.update_data(lon=lon)
    await state.set_state(ObjectFSM.radius)
    await msg.answer("Введи радиус в метрах (например 200):")

@router.message(ObjectFSM.radius)
async def admin_add_object_radius(msg: Message, state: FSMContext):
    u = await db_user_ensure(msg.from_user.id)
    if u["role"] != Role.ADMIN:
        return

    try:
        radius = int((msg.text or "").strip())
        radius = max(50, min(radius, 5000))
    except ValueError:
        await msg.answer("Не похоже на число. Введи радиус ещё раз:")
        return

    data = await state.get_data()
    await state.clear()

    await db_object_add(data["name"], data["lat"], data["lon"], radius)
    await msg.answer(f"✅ Объект добавлен: {data['name']} (r={radius}м)", reply_markup=main_kb(u["role"]))

# =========================
# GEO CHECK
# =========================

async def geo_check_for_today(tg_id: int, lat: float, lon: float) -> Tuple[bool, str]:
    s = await db_shift_get(tg_id, today())
    if not s or not s["object_id"]:
        return False, "Сначала выбери объект: 🏗 Выбрать объект"

    obj = await db_object_get(int(s["object_id"]))
    if not obj:
        return False, "Объект не найден. Выбери заново."

    dist = haversine_m(lat, lon, float(obj["lat"]), float(obj["lon"]))
    ok = dist <= int(obj["radius_m"])
    if ok:
        return True, f"✅ Гео ок (≈{int(dist)}м от '{obj['name']}')"
    return False, f"❌ Ты слишком далеко от '{obj['name']}' (≈{int(dist)}м, нужно ≤ {obj['radius_m']}м)"

# =========================
# WORKDAY FLOW
# =========================

@router.message(F.text == "✅ Пришел")
async def arrived(msg: Message):
    u = await db_user_ensure(msg.from_user.id)
    if not require_approved(u):
        return

    s = await db_shift_get(msg.from_user.id, today())
    if s and s["status"] == DayStatus.SICK:
        await msg.answer("Сегодня отмечено 🤒 Болел. Если это ошибка — напиши админу или снимем отметку.")
        return

    # требуем объект
    if not s or not s["object_id"]:
        await msg.answer("Сначала выбери объект: 🏗 Выбрать объект", reply_markup=main_kb(u["role"]))
        return

    if GEO_REQUIRED:
        await msg.answer("📍 Отправь геолокацию для отметки прихода", reply_markup=kb_geo)
        # отметим в FSM через state (без отдельного флага в БД)
        # сохраним что ждём гео для START
        await msg.bot.send_message(msg.from_user.id, "⚙️ После отправки гео я зафиксирую приход.")
        return

    await db_shift_start(msg.from_user.id, today(), now(), None, None)
    await msg.answer("✅ Приход зафиксирован", reply_markup=main_kb(u["role"]))

@router.message(F.text == "⛔ Ушел")
async def left(msg: Message):
    u = await db_user_ensure(msg.from_user.id)
    if not require_approved(u):
        return

    s = await db_shift_get(msg.from_user.id, today())
    if not s or not s["start_ts"]:
        await msg.answer("Сначала отметь приход: ✅ Пришел")
        return
    if s["status"] == DayStatus.SICK:
        await msg.answer("Сегодня отмечено 🤒 Болел.")
        return

    if GEO_REQUIRED:
        await msg.answer("📍 Отправь геолокацию для отметки ухода", reply_markup=kb_geo)
        await msg.bot.send_message(msg.from_user.id, "⚙️ После отправки гео я зафиксирую уход и часы.")
        return

    end_ts = now()
    hours = calc_hours(s["start_ts"].astimezone(TZ), end_ts)
    await db_shift_end(msg.from_user.id, today(), end_ts, None, None, hours)
    await msg.answer(f"⛔ Уход зафиксирован. Часы за сегодня: **{hours}**", reply_markup=main_kb(u["role"]), parse_mode="Markdown")

@router.message(F.location)
async def got_location(msg: Message):
    u = await db_user_ensure(msg.from_user.id)
    if not require_approved(u):
        return

    lat = float(msg.location.latitude)
    lon = float(msg.location.longitude)

    if GEO_REQUIRED:
        ok, reason = await geo_check_for_today(msg.from_user.id, lat, lon)
        if not ok:
            await msg.answer(reason, reply_markup=main_kb(u["role"]))
            return

    s = await db_shift_get(msg.from_user.id, today())
    if not s:
        await msg.answer("Сначала выбери объект: 🏗 Выбрать объект", reply_markup=main_kb(u["role"]))
        return

    # Если нет start_ts -> это приход, иначе -> уход
    if not s["start_ts"]:
        await db_shift_start(msg.from_user.id, today(), now(), lat, lon)
        await msg.answer(f"{reason if GEO_REQUIRED else '✅'} Приход зафиксирован", reply_markup=main_kb(u["role"]))
        return

    if s["end_ts"]:
        await msg.answer("Уход уже фиксировался сегодня.", reply_markup=main_kb(u["role"]))
        return

    end_ts = now()
    hours = calc_hours(s["start_ts"].astimezone(TZ), end_ts)
    await db_shift_end(msg.from_user.id, today(), end_ts, lat, lon, hours)
    await msg.answer(
        f"{reason if GEO_REQUIRED else '✅'} Уход зафиксирован. Часы за сегодня: **{hours}**",
        reply_markup=main_kb(u["role"]),
        parse_mode="Markdown",
    )

@router.message(F.text == "🤒 Болел")
async def sick(msg: Message):
    u = await db_user_ensure(msg.from_user.id)
    if not require_approved(u):
        return

    await db_shift_set_sick(msg.from_user.id, today(), True)
    await msg.answer("🤒 Отмечено: Болел (сегодня). В табеле будет 'Б'.", reply_markup=main_kb(u["role"]))

@router.message(F.text == "ℹ️ Мой статус")
async def my_status(msg: Message):
    u = await db_user_ensure(msg.from_user.id)
    if not require_approved(u):
        return

    s = await db_shift_get(msg.from_user.id, today())
    if not s:
        await msg.answer("На сегодня записей нет. Выбери объект и отметь приход.", reply_markup=main_kb(u["role"]))
        return

    obj_name = "—"
    if s["object_id"]:
        obj = await db_object_get(int(s["object_id"]))
        if obj:
            obj_name = obj["name"]

    if s["status"] == DayStatus.SICK:
        await msg.answer(f"Сегодня: 🤒 Болел\nОбъект: {obj_name}", reply_markup=main_kb(u["role"]))
        return

    start = s["start_ts"].astimezone(TZ).strftime("%H:%M") if s["start_ts"] else "—"
    end = s["end_ts"].astimezone(TZ).strftime("%H:%M") if s["end_ts"] else "—"
    hours = s["hours"] if s["hours"] is not None else "—"

    await msg.answer(
        f"📅 Сегодня\n🏗 Объект: {obj_name}\n✅ Пришел: {start}\n⛔ Ушел: {end}\n⏱ Часы: {hours}",
        reply_markup=main_kb(u["role"])
    )

# =========================
# ADMIN: EMPLOYEES
# =========================

@router.message(F.text == "📋 Сотрудники")
async def admin_employees(msg: Message):
    u = await db_user_ensure(msg.from_user.id)
    if u["role"] != Role.ADMIN:
        return

    emps = await db_users_list_approved()
    lines = ["👷 Сотрудники (одобренные):", ""]
    for e in emps:
        fio = e["fio"] or f"tg:{e['tg_id']}"
        pos = e["position"] or "—"
        lines.append(f"- {fio} ({pos})")
    await msg.answer("\n".join(lines), reply_markup=main_kb(u["role"]))

# =========================
# XLSX TIMESHEET
# =========================

def build_timesheet_xlsx(year: int, month: int, users: List[asyncpg.Record], shifts: List[asyncpg.Record]) -> str:
    last_day = monthrange(year, month)[1]

    # Map: (tg_id, day) -> (status, hours, object_id)
    by_key = {}
    for s in shifts:
        by_key[(int(s["tg_id"]), int(s["day"].day))] = s

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    # header
    ws["A1"] = "ФИО"
    ws["B1"] = "Должность"
    ws["C1"] = "Объект (день)"
    # days start at col D
    start_col = 4
    for d in range(1, last_day + 1):
        ws.cell(row=1, column=start_col + (d - 1), value=d)

    col_total_hours = start_col + last_day
    col_days_worked = col_total_hours + 1
    col_sick_days = col_total_hours + 2

    ws.cell(row=1, column=col_total_hours, value="Итого часов")
    ws.cell(row=1, column=col_days_worked, value="Отраб. дней")
    ws.cell(row=1, column=col_sick_days, value="Больн. дней")

    header_font = Font(bold=True)
    for c in range(1, col_sick_days + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")

    # objects lookup for day-object note
    # (чтобы быстро подписывать объект в колонке C)
    obj_name_cache = {}

    # rows
    r = 2
    for u in users:
        tg_id = int(u["tg_id"])
        fio = u["fio"] or f"tg:{tg_id}"
        pos = u["position"] or "—"
        ws.cell(row=r, column=1, value=fio)
        ws.cell(row=r, column=2, value=pos)

        total_hours = 0.0
        worked_days = 0
        sick_days = 0
        objects_used = set()

        for d in range(1, last_day + 1):
            s = by_key.get((tg_id, d))
            cell = ws.cell(row=r, column=start_col + (d - 1))

            if not s:
                cell.value = ""
                continue

            if s["status"] == DayStatus.SICK:
                cell.value = "Б"
                sick_days += 1
                continue

            h = float(s["hours"]) if s["hours"] is not None else 0.0
            if h > 0:
                # по ТЗ: ставим 1..8, округлено вверх, не больше 8
                cell.value = h
                total_hours += h
                worked_days += 1
            else:
                cell.value = ""

            if s["object_id"]:
                objects_used.add(int(s["object_id"]))

        # колонка "Объект (день)" — если несколько объектов за месяц, просто перечислим
        # (на практике объект выбирается на день, но для табеля часто достаточно списка)
        if objects_used:
            # оставим ID-шники, чтобы не тянуть БД синхронно тут; названия подставим позже при генерации отчёта (ниже)
            ws.cell(row=r, column=3, value=", ".join([f"#{oid}" for oid in sorted(objects_used)]))
        else:
            ws.cell(row=r, column=3, value="—")

        ws.cell(row=r, column=col_total_hours, value=round(total_hours, 2))
        ws.cell(row=r, column=col_days_worked, value=worked_days)
        ws.cell(row=r, column=col_sick_days, value=sick_days)

        r += 1

    # sizing
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    for c in range(start_col, col_sick_days + 1):
        ws.column_dimensions[get_column_letter(c)].width = 4.5

    # save temp
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path

async def enrich_object_names_in_xlsx(path: str):
    # заменяем в колонке C "#id" на имена объектов
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active

        objs = await db_objects_list()
        m = {int(o["id"]): o["name"] for o in objs}

        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=3).value
            if not v or "#" not in str(v):
                continue
            parts = [p.strip() for p in str(v).split(",")]
            names = []
            for p in parts:
                if p.startswith("#") and p[1:].isdigit():
                    oid = int(p[1:])
                    names.append(m.get(oid, p))
                else:
                    names.append(p)
            ws.cell(row=row, column=3).value = ", ".join(names)

        wb.save(path)
    except Exception:
        logger.exception("Failed to enrich object names in xlsx")

async def send_email_with_attachment(subject: str, body: str, to_emails: List[str], file_path: str):
    if not (SMTP_HOST and SMTP_USER and SMTP_PASS and to_emails):
        logger.info("SMTP not configured or REPORT_EMAILS empty - skip email sending")
        return

    msg = EmailMessage()
    msg["From"] = SMTP_USER
    msg["To"] = ", ".join(to_emails)
    msg["Subject"] = subject
    msg.set_content(body)

    with open(file_path, "rb") as f:
        data = f.read()
    filename = os.path.basename(file_path)
    msg.add_attachment(data, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)

    await aiosmtplib.send(
        msg,
        hostname=SMTP_HOST,
        port=SMTP_PORT,
        username=SMTP_USER,
        password=SMTP_PASS,
        use_tls=True if SMTP_PORT == 465 else False,
        start_tls=True if SMTP_PORT in (587,) else False,
    )

async def generate_and_send_timesheet(bot: Bot, year: int, month: int, send_to_admins: bool = True, send_email: bool = True):
    emps = await db_users_list_approved()
    shifts = await db_month_shifts(year, month)

    path = build_timesheet_xlsx(year, month, emps, shifts)
    await enrich_object_names_in_xlsx(path)

    caption = f"📤 Табель за {year}-{month:02d}"

    # TG admins
    if send_to_admins:
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_document(admin_id, FSInputFile(path), caption=caption)
            except Exception:
                logger.exception("Failed to send timesheet to admin %s", admin_id)

    # Email
    if send_email:
        try:
            await send_email_with_attachment(
                subject=caption,
                body="Табель во вложении.",
                to_emails=REPORT_EMAILS,
                file_path=path
            )
        except Exception:
            logger.exception("Failed to send email timesheet")

    try:
        os.remove(path)
    except Exception:
        pass

@router.message(F.text == "📤 Сформировать табель")
async def admin_make_timesheet(msg: Message):
    u = await db_user_ensure(msg.from_user.id)
    if u["role"] != Role.ADMIN:
        return
    y, m = today().year, today().month
    await generate_and_send_timesheet(msg.bot, y, m, send_to_admins=True, send_email=True)
    await msg.answer("✅ Табель сформирован и отправлен.", reply_markup=main_kb(u["role"]))

# =========================
# SCHEDULER
# =========================

async def scheduled_timesheet_job(bot: Bot):
    d = today()
    last = monthrange(d.year, d.month)[1]
    if d.day in (15, last):
        await generate_and_send_timesheet(bot, d.year, d.month, send_to_admins=True, send_email=True)

# (опционально) автозакрытие смены в 23:59 если забыли "Ушел"
async def scheduled_autoclose_job(bot: Bot):
    d = today()
    # закрываем все открытые смены end_ts=NULL, start_ts NOT NULL, status=work
    async with pool.acquire() as con:
        rows = await con.fetch("""
            SELECT tg_id, start_ts
            FROM shifts
            WHERE day=$1 AND status='work' AND start_ts IS NOT NULL AND end_ts IS NULL
        """, d)

    end_ts = datetime.combine(d, datetime.max.time()).replace(hour=23, minute=59, second=0, microsecond=0, tzinfo=TZ)

    for r in rows:
        tg_id = int(r["tg_id"])
        start_ts = r["start_ts"].astimezone(TZ)
        hours = calc_hours(start_ts, end_ts)
        await db_shift_end(tg_id, d, end_ts, None, None, hours)
        try:
            await bot.send_message(tg_id, f"⏱ Автозакрытие дня: поставил уход 23:59. Часы: {hours}")
        except Exception:
            pass

# =========================
# MAIN
# =========================

async def main():
    if not API_TOKEN:
        raise RuntimeError("API_TOKEN not set")

    await db_init()
    logger.info("Bot starting | GEO_REQUIRED=%s | TZ=%s", GEO_REQUIRED, TZ)

    bot = Bot(API_TOKEN)
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)

    scheduler = AsyncIOScheduler(timezone=str(TZ))

    # Табель: 15 и последний день месяца, 18:00
    scheduler.add_job(scheduled_timesheet_job, "cron", hour=18, minute=0, args=[bot])

    # Автозакрытие дня: 23:59
    scheduler.add_job(scheduled_autoclose_job, "cron", hour=23, minute=59, args=[bot])

    scheduler.start()
    await dp.start_polling(bot)

@router.message()
async def debug_all(msg: Message):
    logger.error(f"DEBUG MESSAGE: {msg.text}")
    await msg.answer("DEBUG: я тебя вижу")

if __name__ == "__main__":
    asyncio.run(main())


