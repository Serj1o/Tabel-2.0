#!/usr/bin/env python3

import asyncio
import os
import logging
from db_logger import DatabaseLogHandler
from datetime import datetime, timedelta, date, time, timezone
from calendar import monthrange
from typing import Dict, Optional
import math
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import holidays
from openpyxl import Workbook
import asyncpg
from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from zoneinfo import ZoneInfo

MOSCOW_TZ = ZoneInfo("Europe/Moscow")

# Настройка
load_dotenv()
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class Config:
    BOT_TOKEN = os.getenv("BOT_TOKEN", "")
    ADMIN_IDS = [int(x.strip()) for x in os.getenv("ADMIN_IDS", "467500951").split(",")]
    DATABASE_URL = os.getenv("DATABASE_URL", "")
    GEO_REQUIRED = os.getenv("GEO_REQUIRED", "true").lower() == "true"
    LOCATION_RADIUS = int(os.getenv("LOCATION_RADIUS", "1000"))
    LOCATION_MAX_AGE_SECONDS = int(os.getenv("LOCATION_MAX_AGE_SECONDS", "300"))
    LOCATION_MAX_ACCURACY = float(os.getenv("LOCATION_MAX_ACCURACY", "250"))
    SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
    SMTP_USERNAME = os.getenv("SMTP_USERNAME", "")
    SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
    EMAIL_RECIPIENTS = [x.strip() for x in os.getenv("EMAIL_RECIPIENTS", "").split(",") if x.strip()]
    WORK_START_HOUR = 9
    MAX_WORK_HOURS = 8

MOSCOW_TZ = ZoneInfo("Europe/Moscow")


def now_moscow() -> datetime:
    """Текущее московское время с информацией о часовом поясе."""
    return datetime.now(MOSCOW_TZ)


def now_moscow_naive() -> datetime:
    """Текущее московское время без tzinfo для сохранения в БД."""
    return now_moscow().replace(tzinfo=None)


def today_moscow() -> date:
    """Текущая дата по московскому времени."""
    return now_moscow().date()

if not Config.BOT_TOKEN or not Config.DATABASE_URL:
    raise ValueError("Проверьте BOT_TOKEN и DATABASE_URL в переменных окружения")

if Config.DATABASE_URL:
    logger.info(f"Database URL получен, длина: {len(Config.DATABASE_URL)} символов")
    # Проверяем формат PostgreSQL URL
    if not Config.DATABASE_URL.startswith(('postgresql://', 'postgres://')):
        logger.warning("Database URL может иметь неверный формат. Ожидается postgresql:// или postgres://")
else:
    logger.error("DATABASE_URL не установлен!")
    
# Простая функция расчета расстояния (в километрах)
def calculate_distance(lat1, lon1, lat2, lon2):
    """Упрощенный расчет расстояния между двумя точками (в метрах)"""
    if not all([lat1, lon1, lat2, lon2]):
        return float('inf')
    
    # Преобразуем градусы в радианы
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    
    # Формула гаверсинусов
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    r = 6371000  # Радиус Земли в метрах
    
    return c * r

# Состояния
class Form(StatesGroup):
    waiting_for_location = State()
    waiting_for_object = State()
    waiting_for_employee_name = State()
    waiting_for_position = State()
    waiting_for_object_data = State()
    waiting_for_sick_reason = State()
    waiting_for_employee_admin_data = State()
    waiting_for_manual_note = State()
    waiting_for_manual_photo = State()

# Основной класс бота
class WorkTimeBot:
    def __init__(self):
        self.bot = Bot(token=Config.BOT_TOKEN)
        self.dp = Dispatcher(storage=MemoryStorage())
        self.scheduler = AsyncIOScheduler(timezone="Europe/Moscow")
        self.pool = None
        self.db_log_handler: Optional[DatabaseLogHandler] = None
        self._holiday_cache: Dict[int, holidays.HolidayBase] = {}
        self.register_handlers()

    async def init_db(self):
        """Инициализация БД"""
        try:
            self.pool = await asyncpg.create_pool(Config.DATABASE_URL)

            async with self.pool.acquire() as conn:
                # Таблица сотрудников - ДОБАВЛЕНА КОЛОНКА position
                await conn.execute('''
                    CREATE TABLE IF NOT EXISTS employees (
                        id SERIAL PRIMARY KEY,
                        telegram_id BIGINT UNIQUE,
                        full_name VARCHAR(255) NOT NULL,
                        position VARCHAR(100),  -- НОВАЯ КОЛОНКА
                        is_admin BOOLEAN DEFAULT FALSE,
                        is_active BOOLEAN DEFAULT TRUE,
                        is_approved BOOLEAN DEFAULT FALSE
                    )
                ''')

                # Таблица объектов
                await conn.execute('''
                    CREATE TABLE IF NOT EXISTS objects (
                        id SERIAL PRIMARY KEY,
                        name VARCHAR(255) NOT NULL,
                        address TEXT,
                        latitude DECIMAL(10, 8),
                        longitude DECIMAL(11, 8),
                        radius INTEGER DEFAULT 500
                    )
                ''')

                # Таблица рабочих отметок
                await conn.execute('''
                    CREATE TABLE IF NOT EXISTS time_logs (
                        id SERIAL PRIMARY KEY,
                        employee_id INTEGER REFERENCES employees(id),
                        object_id INTEGER REFERENCES objects(id),
                        date DATE NOT NULL,
                        check_in TIMESTAMP,
                        check_out TIMESTAMP,
                        check_in_lat DECIMAL(10, 8),
                        check_in_lon DECIMAL(11, 8),
                        check_out_lat DECIMAL(10, 8),
                        check_out_lon DECIMAL(11, 8),
                        hours_worked DECIMAL(4, 2) DEFAULT 0,
                        status VARCHAR(20) DEFAULT 'work',
                        notes TEXT
                    )
                ''')

                # Таблица запросов на доступ
                await conn.execute('''
                    CREATE TABLE IF NOT EXISTS access_requests (
                        id SERIAL PRIMARY KEY,
                        telegram_id BIGINT NOT NULL,
                        full_name VARCHAR(255),
                        position VARCHAR(100),  -- НОВАЯ КОЛОНКА
                        status VARCHAR(20) DEFAULT 'pending'
                    )
                ''')

                await conn.execute('''
                    CREATE TABLE IF NOT EXISTS log_entries (
                        id SERIAL PRIMARY KEY,
                        created_at TIMESTAMP DEFAULT NOW(),
                        logger VARCHAR(255),
                        level VARCHAR(50),
                        message TEXT
                    )
                ''')

                await conn.execute('''
                    CREATE TABLE IF NOT EXISTS location_audit (
                        id SERIAL PRIMARY KEY,
                        employee_id INTEGER REFERENCES employees(id),
                        object_id INTEGER REFERENCES objects(id),
                        action VARCHAR(20),
                        latitude DECIMAL(10, 8),
                        longitude DECIMAL(11, 8),
                        distance INTEGER,
                        accuracy DECIMAL(8, 2),
                        message_time TIMESTAMP,
                        is_suspicious BOOLEAN DEFAULT FALSE,
                        reason TEXT,
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                ''')

                # Гарантируем наличие критичных столбцов для обратной совместимости
                await self.safe_migration()

                # Добавляем администратора БЕЗ ДОЛЖНОСТИ (чтобы не попадал в табель)
                await conn.execute('''
                    INSERT INTO employees (telegram_id, full_name, is_admin, is_approved)
                    VALUES ($1, $2, $3, $4)
                    ON CONFLICT (telegram_id) DO UPDATE SET
                    is_admin = EXCLUDED.is_admin,
                    is_approved = EXCLUDED.is_approved
                ''', Config.ADMIN_IDS[0], "Главный Администратор", True, True)

                self.attach_db_logger()

                logger.info("База данных инициализирована")

        except Exception as e:
            logger.error(f"Ошибка при инициализации БД: {e}")
            raise

    async def ensure_column(self, conn: asyncpg.Connection, table: str, column: str, definition: str):
        """Добавляет столбец, если его нет (для миграций без простоя)"""
        try:
            column_exists = await conn.fetchval(
                """
                SELECT 1 FROM information_schema.columns 
                WHERE table_name = $1 AND column_name = $2
                """,
                table,
                column.lower(),
            )

            if not column_exists:
                await conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
                logger.info(f"Добавлен столбец {column} в таблицу {table}")
            else:
                logger.info(f"Столбец {column} уже существует в таблице {table}")
        except Exception as err:
            logger.error(f"Ошибка при проверке/добавлении столбца {column} в таблицу {table}: {err}")
            try:
                await conn.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {definition}")
                logger.info(f"Добавлен столбец {column} в таблицу {table} (использован IF NOT EXISTS)")
            except Exception as err2:
                logger.error(f"Критическая ошибка добавления столбца: {err2}")
                raise

    async def ensure_position_columns(self, conn: asyncpg.Connection):
        """Гарантирует наличие столбца position в ключевых таблицах."""
        await conn.execute("CREATE TABLE IF NOT EXISTS employees (id SERIAL PRIMARY KEY)")
        await conn.execute("CREATE TABLE IF NOT EXISTS access_requests (id SERIAL PRIMARY KEY)")

        await self.ensure_column(conn, "employees", "position", "VARCHAR(100)")
        await self.ensure_column(conn, "access_requests", "position", "VARCHAR(100)")

    async def safe_migration(self):
        """Безопасная миграция БД при старте"""
        try:
            async with self.pool.acquire() as conn:
                await self.ensure_position_columns(conn)
                await conn.execute('''
                    CREATE TABLE IF NOT EXISTS log_entries (
                        id SERIAL PRIMARY KEY,
                        created_at TIMESTAMP DEFAULT NOW(),
                        logger VARCHAR(255),
                        level VARCHAR(50),
                        message TEXT
                    )
                ''')
                await conn.execute('''
                    CREATE TABLE IF NOT EXISTS location_audit (
                        id SERIAL PRIMARY KEY,
                        employee_id INTEGER REFERENCES employees(id),
                        object_id INTEGER REFERENCES objects(id),
                        action VARCHAR(20),
                        latitude DECIMAL(10, 8),
                        longitude DECIMAL(11, 8),
                        distance INTEGER,
                        accuracy DECIMAL(8, 2),
                        message_time TIMESTAMP,
                        is_suspicious BOOLEAN DEFAULT FALSE,
                        reason TEXT,
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                ''')

            logger.info("Миграция БД завершена успешно")
        except Exception as e:
            logger.error(f"Ошибка миграции БД: {e}")
            raise

    def attach_db_logger(self):
        """Подключает обработчик логов, сохраняющий записи в БД"""
        if not self.pool or self.db_log_handler:
            return

        self.db_log_handler = DatabaseLogHandler(self.pool)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        self.db_log_handler.setFormatter(formatter)
        logging.getLogger().addHandler(self.db_log_handler)

    def moscow_now(self) -> datetime:
        """Возвращает текущее московское время с TZ-информацией"""
        return datetime.now(MOSCOW_TZ)

    def moscow_now_naive(self) -> datetime:
        """Возвращает текущее московское время без TZ (для хранения в БД без таймзоны)"""
        return self.moscow_now().replace(tzinfo=None)

    def moscow_today(self) -> date:
        """Текущая дата в московском часовом поясе"""
        return self.moscow_now().date()

    def _get_ru_holidays(self, year: int) -> holidays.HolidayBase:
        """Кэширует календарь праздничных дней РФ"""
        if year not in self._holiday_cache:
            self._holiday_cache[year] = holidays.Russia(years=[year])
        return self._holiday_cache[year]

    def is_working_day(self, target_date: date) -> bool:
        """Проверяет, является ли дата рабочим днем с учетом праздников РФ"""
        return target_date.weekday() < 5 and target_date not in self._get_ru_holidays(target_date.year)

    def normalize_message_time(self, msg_time: datetime) -> datetime:
        """Приводит время сообщения к московскому часовому поясу"""
        if msg_time.tzinfo is None:
            msg_time = msg_time.replace(tzinfo=timezone.utc)
        return msg_time.astimezone(MOSCOW_TZ)

    async def build_main_keyboard_for_user(self, user_id: int) -> ReplyKeyboardMarkup:
        """Формирует главное меню с учетом роли и статуса больничного на сегодня"""
        is_admin = False
        sick_today = False

        async with self.pool.acquire() as conn:
            is_admin = bool(await conn.fetchval(
                'SELECT is_admin FROM employees WHERE telegram_id = $1',
                user_id,
            ) or False)
            sick_today = bool(await conn.fetchval(
                '''
                SELECT 1 FROM time_logs
                WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1)
                  AND date = $2 AND status = 'sick'
                LIMIT 1
                ''',
                user_id,
                self.moscow_today(),
            ) or False)

        return self.get_main_keyboard(is_admin=is_admin, show_remove_sick=sick_today)
     
    def get_main_keyboard(self, is_admin: bool = False, show_remove_sick: bool = False) -> ReplyKeyboardMarkup:
        """Создает основное меню с кнопками"""
        builder = ReplyKeyboardBuilder()

        if is_admin:
            builder.add(KeyboardButton(text="Статистика"))
            builder.add(KeyboardButton(text="Админ панель"))
            builder.adjust(1)
            return builder.as_markup(resize_keyboard=True)

        # Основные кнопки для сотрудников
        builder.add(KeyboardButton(text="Пришел"))
        builder.add(KeyboardButton(text="Ушел"))
        builder.add(KeyboardButton(text="Болел"))
        if show_remove_sick:
            builder.add(KeyboardButton(text="Снять больничный"))
        builder.add(KeyboardButton(text="Выбрать объект"))
        builder.add(KeyboardButton(text="Отправить геолокацию", request_location=True))

        builder.adjust(2)  # 2 кнопки в ряд
        return builder.as_markup(resize_keyboard=True)
    
    def register_handlers(self):
        """Регистрация обработчиков"""
        # Команды
        self.dp.message.register(self.handle_start, Command("start"))
        self.dp.message.register(self.handle_admin, Command("admin"))

        # Отмена в любом состоянии
        self.dp.message.register(self.handle_cancel, F.text == "Отмена")

        # Обработчики кнопок меню
        self.dp.message.register(self.handle_come, F.text == "Пришел")
        self.dp.message.register(self.handle_leave, F.text == "Ушел")
        self.dp.message.register(self.handle_sick_btn, F.text == "Болел")
        self.dp.message.register(self.handle_sick_clear, F.text == "Снять больничный")
        self.dp.message.register(self.handle_select_object_btn, F.text == "Выбрать объект")
        self.dp.message.register(self.handle_stats_btn, F.text == "Статистика")
        self.dp.message.register(self.handle_admin_btn, F.text == "Админ панель")
        
        # Геолокация
        self.dp.message.register(self.handle_location, F.location)
        
        # Состояния
        self.dp.message.register(self.process_text, Form.waiting_for_employee_name)
        self.dp.message.register(self.process_object_data, Form.waiting_for_object_data)
        self.dp.message.register(self.process_sick_reason, Form.waiting_for_sick_reason)
        self.dp.message.register(self.process_employee_admin_data, Form.waiting_for_employee_admin_data)
        self.dp.message.register(self.process_manual_note, Form.waiting_for_manual_note)
        self.dp.message.register(self.process_manual_photo, Form.waiting_for_manual_photo)
        
        # Callback-запросы
        self.dp.callback_query.register(
            self.handle_callback,
            F.data.startswith("obj_")
            | F.data.startswith("admin_")
            | F.data.startswith("approve_")
            | F.data.startswith("reject_")
            | F.data.startswith("sick_")
            | (F.data == "request_access"),
        )
    
    # Вспомогательные методы
    async def check_access(self, user_id: int, need_admin: bool = False) -> bool:
        """Проверка доступа пользователя"""
        if not self.pool:
            return False
        
        async with self.pool.acquire() as conn:
            query = '''
                SELECT is_approved, is_admin FROM employees 
                WHERE telegram_id = $1 AND is_active = TRUE
            '''
            user = await conn.fetchrow(query, user_id)
            
            if not user:
                return False
            
            if need_admin:
                return user['is_admin'] and user['is_approved']
            
            return user['is_approved']
    
    # Основные обработчики
    async def handle_start(self, message: types.Message, state: FSMContext):
        """Обработка /start"""
        user_id = message.from_user.id
        
        async with self.pool.acquire() as conn:
            user = await conn.fetchrow('''
                SELECT full_name, is_admin, is_approved FROM employees 
                WHERE telegram_id = $1
            ''', user_id)
        
        if user and user['is_approved']:
            text = f"Добро пожаловать, {user['full_name']}!\n\n"
            if user['is_admin']:
                text += "Вы администратор системы.\n"
            text += "Используйте меню ниже для работы с системой."
            
            keyboard = await self.build_main_keyboard_for_user(user_id)
            await message.answer(text, reply_markup=keyboard)
        else:
            keyboard = InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="Запросить доступ", callback_data="request_access")
            ]])
            await message.answer("Это закрытая система учета рабочего времени. Запросите доступ:",
                               reply_markup=keyboard)

    async def handle_cancel(self, message: types.Message, state: FSMContext):
        """Обработка кнопки 'Отмена' из любого меню"""
        await self.return_to_main_menu(message, state)
    
    async def handle_come(self, message: types.Message, state: FSMContext):
        """Обработка кнопки 'Пришел'"""
        user_id = message.from_user.id

        if not await self.check_access(user_id):
            await message.answer("Доступ запрещен.")
            return

        today = self.moscow_today()
        async with self.pool.acquire() as conn:
            user = await conn.fetchrow(
                'SELECT is_admin FROM employees WHERE telegram_id = $1',
                user_id,
            )
            if user and user['is_admin']:
                await message.answer(
                    "Отметки прихода недоступны администраторам.",
                    reply_markup=self.get_main_keyboard(is_admin=True),
                )
                return

            sick_today = await conn.fetchval(
                '''
                SELECT 1 FROM time_logs
                WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1)
                  AND date = $2 AND status = 'sick'
                LIMIT 1
                ''',
                user_id,
                today,
            )
            existing = await conn.fetchval('''
                SELECT check_in FROM time_logs
                WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1) AND date = $2
                LIMIT 1
            ''', user_id, today)

        if sick_today:
            await message.answer(
                "Нельзя отметить приход в день, когда стоит статус 'Болел'.",
                reply_markup=await self.build_main_keyboard_for_user(user_id),
            )
            return

        if existing:
            await message.answer(f"Вы уже отметили приход в {existing.strftime('%H:%M')}")
            return

        await state.set_state(Form.waiting_for_location)
        await state.update_data(action="checkin")
        await self.show_objects(message, state, "checkin")
    
    async def handle_leave(self, message: types.Message, state: FSMContext):
        """Обработка кнопки 'Ушел'"""
        user_id = message.from_user.id

        if not await self.check_access(user_id):
            await message.answer("Доступ запрещен.")
            return

        today = self.moscow_today()
        async with self.pool.acquire() as conn:
            user = await conn.fetchrow(
                'SELECT is_admin FROM employees WHERE telegram_id = $1',
                user_id,
            )
            if user and user['is_admin']:
                await message.answer(
                    "Отметки ухода недоступны администраторам.",
                    reply_markup=self.get_main_keyboard(is_admin=True),
                )
                return

            sick_today = await conn.fetchval(
                '''
                SELECT 1 FROM time_logs
                WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1)
                  AND date = $2 AND status = 'sick'
                LIMIT 1
                ''',
                user_id,
                today,
            )
            log = await conn.fetchrow('''
                SELECT id, check_in FROM time_logs
                WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1) AND date = $2
                AND check_out IS NULL
                LIMIT 1
            ''', user_id, today)

        if sick_today:
            await message.answer(
                "Статус 'Болел' отмечен на сегодня. Отметка ухода недоступна.",
                reply_markup=await self.build_main_keyboard_for_user(user_id),
            )
            await state.clear()
            return

        if not log:
            await message.answer(
                "Сначала отметьте приход.",
                reply_markup=self.get_main_keyboard(
                    is_admin=await self.check_access(user_id, need_admin=True)
                ),
            )
            await state.clear()
            return

        await state.set_state(Form.waiting_for_location)
        await state.update_data(action="checkout", log_id=log['id'])
        await self.show_objects(message, state, "checkout")

    async def return_to_main_menu(self, message: types.Message, state: FSMContext):
        """Возврат в главное меню"""
        user_id = message.from_user.id
        async with self.pool.acquire() as conn:
            keyboard = await self.build_main_keyboard_for_user(user_id)
        await message.answer("Операция отменена", reply_markup=keyboard)
        await state.clear()
    
    async def handle_sick_btn(self, message: types.Message, state: FSMContext):
        """Обработка кнопки 'Болел'"""
        user_id = message.from_user.id

        if not await self.check_access(user_id):
            await message.answer("Доступ запрещен.")
            return

        today = self.moscow_today()
        async with self.pool.acquire() as conn:
            working_today = await conn.fetchval(
                '''
                SELECT 1 FROM time_logs
                WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1)
                  AND date = $2 AND check_in IS NOT NULL
                LIMIT 1
                ''',
                user_id,
                today,
            )
            sick_today = await conn.fetchval(
                '''
                SELECT 1 FROM time_logs
                WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1)
                  AND date = $2 AND status = 'sick'
                LIMIT 1
                ''',
                user_id,
                today,
            )

        if working_today:
            await message.answer(
                "Нельзя поставить статус 'Болел', пока отмечен приход.",
                reply_markup=await self.build_main_keyboard_for_user(user_id),
            )
            return
        if sick_today:
            await message.answer(
                "Болничный уже отмечен на сегодня.",
                reply_markup=await self.build_main_keyboard_for_user(user_id),
            )
            return

        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Подтвердить больничный", callback_data="sick_confirm")],
            [InlineKeyboardButton(text="Отмена", callback_data="sick_cancel")],
        ])
        await message.answer(
            "Отметить день как больничный?",
            reply_markup=keyboard,
        )
        await state.update_data(user_id=user_id)
    
    async def handle_select_object_btn(self, message: types.Message, state: FSMContext):
        """Обработка кнопки 'Выбрать объект'"""
        user_id = message.from_user.id

        if not await self.check_access(user_id):
            await message.answer("Доступ запрещен.")
            return

        async with self.pool.acquire() as conn:
            is_admin = await conn.fetchval('SELECT is_admin FROM employees WHERE telegram_id = $1', user_id)
            if is_admin:
                await message.answer(
                    "Выбор объекта доступен только сотрудникам.",
                    reply_markup=self.get_main_keyboard(is_admin=True),
                )
                return

        await self.show_objects(message, state, "select")
    
    async def handle_stats_btn(self, message: types.Message):
        """Обработка кнопки 'Статистика'"""
        user_id = message.from_user.id

        if not await self.check_access(user_id, need_admin=True):
            await message.answer("Доступно только администраторам.")
            return
        
        today = self.moscow_today()

        async with self.pool.acquire() as conn:
            employees = await conn.fetch(
                '''
                SELECT id, full_name
                FROM employees
                WHERE is_active = TRUE AND is_approved = TRUE AND is_admin = FALSE
                ORDER BY full_name
                '''
            )
            logs = await conn.fetch(
                '''
                SELECT tl.employee_id, tl.check_in, tl.status,
                       COALESCE(o.name, 'Не указан') AS object_name
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                LEFT JOIN objects o ON tl.object_id = o.id
                WHERE tl.date = $1 AND e.is_active = TRUE AND e.is_approved = TRUE
                ORDER BY e.full_name
                ''',
                today,
            )
            missing_checkins = await conn.fetch(
                '''
                SELECT e.full_name
                FROM employees e
                WHERE e.is_active = TRUE AND e.is_approved = TRUE
                  AND NOT EXISTS (
                      SELECT 1 FROM time_logs tl
                      WHERE tl.employee_id = e.id AND tl.date = $1 AND tl.check_in IS NOT NULL
                  )
                  AND NOT EXISTS (
                      SELECT 1 FROM time_logs tl
                      WHERE tl.employee_id = e.id AND tl.date = $1 AND tl.status = 'sick'
                  )
                ORDER BY e.full_name
                ''',
                today,
            )
            missing_checkouts = await conn.fetch(
                '''
                SELECT e.full_name
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                WHERE tl.date = $1 AND tl.check_in IS NOT NULL AND tl.check_out IS NULL
                  AND e.is_active = TRUE AND e.is_approved = TRUE
                ORDER BY e.full_name
                ''',
                today,
            )

        log_map: Dict[int, asyncpg.Record] = {}
        for log in logs:
            existing = log_map.get(log['employee_id'])
            if not existing:
                log_map[log['employee_id']] = log
                continue
            if log['check_in'] and (existing['check_in'] is None or log['check_in'] < existing['check_in']):
                log_map[log['employee_id']] = log

        arrived = []
        sick = []
        missing = []

        for emp in employees:
            entry = log_map.get(emp['id'])
            if not entry:
                missing.append(emp['full_name'])
                continue
            if entry['status'] == 'sick':
                sick.append(emp['full_name'])
                continue
            if entry['check_in']:
                arrived.append(
                    (
                        emp['full_name'],
                        entry['check_in'],
                        entry['object_name'],
                    )
                )
            else:
                missing.append(emp['full_name'])

        text_sections = ["Статистика по персоналу за сегодня:\n"]

        def format_block(title: str, rows: list[str]):
            if not rows:
                return f"{title}: нет\n"
            return f"{title}:\n" + "\n".join(rows) + "\n"

        arrived_rows = [
            f"{idx}. {name} — {check_in.strftime('%H:%M')} ({obj})"
            for idx, (name, check_in, obj) in enumerate(arrived, start=1)
        ]
        missing_rows = [f"{idx}. {name}" for idx, name in enumerate(missing, start=1)]
        sick_rows = [f"{idx}. {name}" for idx, name in enumerate(sick, start=1)]

        text_sections.append(format_block("Пришли", arrived_rows))
        text_sections.append(format_block("Не отметились", missing_rows))
        text_sections.append(format_block("Болеют", sick_rows))

        text += "\nНе отметили приход:\n"
        if not missing_checkins:
            text += "—\n"
        else:
            for row in missing_checkins:
                text += f"• {row['full_name']}\n"

        text += "\nНе отметили уход:\n"
        if not missing_checkouts:
            text += "—"
        else:
            for row in missing_checkouts:
                text += f"• {row['full_name']}\n"

        text += "\nНе отметили приход:\n"
        if not missing_checkins:
            text += "—\n"
        else:
            for row in missing_checkins:
                text += f"• {row['full_name']}\n"

        text += "\nНе отметили уход:\n"
        if not missing_checkouts:
            text += "—"
        else:
            for row in missing_checkouts:
                text += f"• {row['full_name']}\n"

        await message.answer(text)
    
    async def handle_admin_btn(self, message: types.Message):
        """Обработка кнопки 'Админ панель'"""
        await self.handle_admin(message)
    
    async def handle_admin(self, message: types.Message):
        """Панель администратора"""
        user_id = message.from_user.id
        
        if not await self.check_access(user_id, need_admin=True):
            await message.answer("Нет прав администратора.")
            return
        
        keyboard = InlineKeyboardBuilder()
        buttons = [
            ("Запросы на доступ", "admin_requests"),
            ("Сформировать табель", "admin_timesheet"),
            ("Отправить табель", "admin_send"),
            ("Сегодняшние приходы", "admin_today_checkins"),
            ("Список сотрудников", "admin_employees"),
            ("Добавить сотрудника", "admin_add_employee"),
            ("Изменить сотрудника", "admin_edit_employee"),
            ("Удалить сотрудника", "admin_delete_employee"),
            ("Список объектов", "admin_objects"),
            ("Добавить объект", "admin_add_object"),
            ("Изменить объект", "admin_edit_object"),
            ("Удалить объект", "admin_delete_object"),
            ("Статистика", "admin_stats"),
            ("Экспорт действий", "admin_export_logs"),
        ]
        
        for text, data in buttons:
            keyboard.button(text=text, callback_data=data)
        keyboard.adjust(2)

        await message.answer("Панель администратора:", reply_markup=keyboard.as_markup())

    async def handle_callback(self, callback: types.CallbackQuery, state: FSMContext):
        """Обработка callback-запросов"""
        data = callback.data
        
        if data.startswith("obj_"):
            await self.handle_object_selection(callback, state)
        elif data.startswith("approve_"):
            await self.handle_approval(callback)
        elif data.startswith("reject_"):
            await self.handle_rejection(callback)
        elif data.startswith("admin_"):
            if not await self.check_access(callback.from_user.id, need_admin=True):
                await callback.answer("Нет прав", show_alert=True)
                return

            # Подстраховка: перед любыми админ-действиями проверяем наличие критичных столбцов
            await self.safe_migration()

            try:
                if data == "admin_requests":
                    await self.show_pending_requests(callback)  # Используем правильное имя метода
                elif data == "admin_timesheet":
                    await self.ask_timesheet_type(callback)
                elif data == "admin_timesheet_overall":
                    await self.generate_timesheet(callback, include_objects=False)
                elif data == "admin_timesheet_detailed":
                    await self.generate_timesheet(callback, include_objects=True)
                elif data == "admin_send":
                    await self.send_timesheet_email(callback)
                elif data == "admin_today_checkins":
                    await self.show_today_checkins(callback)
                elif data == "admin_employees":
                    await self.show_employees(callback)
                elif data == "admin_add_employee":
                    await self.start_employee_input(callback, state, action="add")
                elif data == "admin_edit_employee":
                    await self.choose_employee_for_action(callback, "edit")
                elif data == "admin_delete_employee":
                    await self.choose_employee_for_action(callback, "delete")
                elif data == "admin_objects":
                    await self.show_objects_admin(callback)
                elif data == "admin_add_object":
                    await self.start_object_input(callback, state, action="add")
                elif data == "admin_edit_object":
                    await self.choose_object_for_action(callback, "edit")
                elif data == "admin_delete_object":
                    await self.choose_object_for_action(callback, "delete")
                elif data.startswith("admin_obj_edit_"):
                    await self.start_object_input(callback, state, action="edit")
                elif data.startswith("admin_obj_delete_"):
                    await self.delete_object(callback)
                elif data.startswith("admin_emp_view_"):
                    await self.show_employee_card(callback)
                elif data.startswith("admin_emp_edit_"):
                    await self.start_employee_input(callback, state, action="edit")
                elif data.startswith("admin_emp_toggle_"):
                    await self.toggle_employee_flag(callback)
                elif data.startswith("admin_emp_delete_"):
                    await self.delete_employee(callback)
                elif data == "admin_stats":
                    await self.show_stats(callback)
                elif data == "admin_export_logs":
                    await self.send_action_log(callback)
                else:
                    await callback.answer("Неизвестная команда", show_alert=True)
                    return
            except Exception as admin_error:
                logger.error(f"Ошибка обработки admin callback {data}: {admin_error}")
                await callback.message.answer("Произошла ошибка при обработке запроса администратора. Попробуйте позже.")
        elif data == "sick_confirm":
            await callback.message.answer(
                "Введите причину больничного (необязательно):",
                reply_markup=types.ReplyKeyboardRemove(),
            )
            await callback.answer()
            await state.set_state(Form.waiting_for_sick_reason)
            return
        elif data == "sick_cancel":
            await self.return_to_main_menu(callback.message, state)
            await callback.answer("Отменено")
            return
        elif data == "request_access":
            await callback.message.answer("Введите ваше ФИО и должность (через |):")
            await state.set_state(Form.waiting_for_employee_name)
            await state.update_data(action="request_access", user_id=callback.from_user.id)

        await callback.answer()

    async def show_pending_requests(self, callback: types.CallbackQuery):
        """Показать ожидающие запросы на доступ"""
        async with self.pool.acquire() as conn:
            await self.ensure_position_columns(conn)
            try:
                requests = await conn.fetch('''
                    SELECT id, telegram_id, full_name, position FROM access_requests
                    WHERE status = 'pending' ORDER BY id DESC
                ''')
            except asyncpg.UndefinedColumnError:
                # Если база еще не мигрирована, мигрируем на лету и пробуем снова
                await self.ensure_position_columns(conn)
                requests = await conn.fetch('''
                    SELECT id, telegram_id, full_name, position FROM access_requests
                    WHERE status = 'pending' ORDER BY id DESC
                ''')
        
        if not requests:
            await callback.message.answer("Нет ожидающих запросов")
            return
        
        for req in requests:
            keyboard = InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="Одобрить", callback_data=f"approve_{req['id']}"),
                InlineKeyboardButton(text="Отклонить", callback_data=f"reject_{req['id']}")
            ]])
            
            position_text = f"\nДолжность: {req['position']}" if req['position'] else ""
            await callback.message.answer(
                f"Запрос #{req['id']}\nФИО: {req['full_name']}{position_text}\nID: {req['telegram_id']}", 
                reply_markup=keyboard
            )
    
    async def show_requests(self, callback: types.CallbackQuery):
        """Показать запросы на доступ - синоним для show_pending_requests"""
        await self.show_pending_requests(callback)

    async def show_objects(self, message: types.Message, state: FSMContext, action: str):
        """Показать список объектов для выбора"""
        async with self.pool.acquire() as conn:
            objects = await conn.fetch('SELECT id, name, address FROM objects ORDER BY name')
        
        if not objects:
            await message.answer("Список объектов пуст. Обратитесь к администратору.",
                               reply_markup=self.get_main_keyboard())
            await state.clear()
            return
        
        keyboard = InlineKeyboardBuilder()
        for obj in objects:
            title = obj['name']
            if obj['address']:
                title += f" ({obj['address']})"
            keyboard.button(text=title, callback_data=f"obj_{obj['id']}_{action}")
        keyboard.adjust(1)

        await state.update_data(action=action)
        await message.answer("Выберите объект:", reply_markup=keyboard.as_markup())

    async def handle_object_selection(self, callback: types.CallbackQuery, state: FSMContext):
        """Обработка выбора объекта"""
        _, obj_id_str, action = callback.data.split("_", maxsplit=2)
        obj_id = int(obj_id_str)

        async with self.pool.acquire() as conn:
            obj = await conn.fetchrow('SELECT id, name, address, latitude, longitude, radius FROM objects WHERE id = $1', obj_id)
        
        if not obj:
            await callback.message.answer("Объект не найден, попробуйте снова.")
            return
        
        data = await state.get_data()
        lat = data.get("lat")
        lon = data.get("lon")
        log_id = data.get("log_id")
        user_id = callback.from_user.id

        if action == "checkin":
            await state.update_data(selected_object_id=obj['id'])
            if Config.GEO_REQUIRED:
                keyboard = ReplyKeyboardMarkup(
                    keyboard=[[KeyboardButton(text="Отправить геолокацию", request_location=True)],
                              [KeyboardButton(text="Отмена")]],
                    resize_keyboard=True,
                )
                await callback.message.answer(
                    f"Объект выбран: {obj['name']}\nОтправьте геолокацию для отметки прихода.",
                    reply_markup=keyboard,
                )
                await state.set_state(Form.waiting_for_location)
            else:
                await self.process_checkin_manual(user_id, obj, lat, lon)
                await state.clear()
        elif action == "checkout":
            if not log_id:
                await callback.message.answer("Не удалось найти запись для отметки ухода. Попробуйте еще раз.")
            else:
                await state.update_data(selected_object_id=obj['id'])
                if Config.GEO_REQUIRED:
                    keyboard = ReplyKeyboardMarkup(
                        keyboard=[[KeyboardButton(text="Отправить геолокацию", request_location=True)],
                                  [KeyboardButton(text="Отмена")]],
                        resize_keyboard=True,
                    )
                    await callback.message.answer(
                        f"Объект выбран: {obj['name']}\nОтправьте геолокацию для отметки ухода.",
                        reply_markup=keyboard,
                    )
                    await state.set_state(Form.waiting_for_location)
                else:
                    await self.process_checkout_manual(log_id, obj, user_id, lat, lon)
                    await state.clear()
        else:
            await state.update_data(selected_object_id=obj['id'])
            await callback.message.answer(
                f"Объект выбран: {obj['name']}\nИспользуйте отметку прихода/ухода для записи времени.",
                reply_markup=self.get_main_keyboard()
            )
    
    # Обработчики состояний
    async def process_text(self, message: types.Message, state: FSMContext):
        """Обработка текстовых сообщений"""
        data = await state.get_data()
        action = data.get("action")
        
        if action == "request_access":
            raw_text = message.text.strip()
            parts = [part.strip() for part in raw_text.split("|", maxsplit=1)]
            full_name = parts[0]
            position = parts[1] if len(parts) > 1 else None
            user_id = data.get("user_id")

            if not full_name:
                await message.answer("Пожалуйста, введите ваше ФИО (можно через '|' указать должность).")
                return

            async with self.pool.acquire() as conn:
                await self.ensure_position_columns(conn)

                try:
                    if position:
                        await conn.execute('''
                            INSERT INTO access_requests (telegram_id, full_name, position)
                            VALUES ($1, $2, $3)
                        ''', user_id, full_name, position)
                    else:
                        await conn.execute('''
                            INSERT INTO access_requests (telegram_id, full_name, position)
                            VALUES ($1, $2, NULL)
                        ''', user_id, full_name)
                except asyncpg.UndefinedColumnError:
                    # Если колонка отсутствует в старой схеме, создаем ее и повторяем вставку
                    await self.ensure_position_columns(conn)
                    if position:
                        await conn.execute('''
                            INSERT INTO access_requests (telegram_id, full_name, position)
                            VALUES ($1, $2, $3)
                        ''', user_id, full_name, position)
                    else:
                        await conn.execute('''
                            INSERT INTO access_requests (telegram_id, full_name, position)
                            VALUES ($1, $2, NULL)
                        ''', user_id, full_name)
                
                # Уведомляем администраторов
                admins = await conn.fetch('SELECT telegram_id FROM employees WHERE is_admin = TRUE')
                for admin in admins:
                    try:
                        position_line = f"\nДолжность: {position}" if position else ""
                        await self.bot.send_message(
                            admin['telegram_id'],
                            f"Новый запрос на доступ:\nФИО: {full_name}{position_line}\nID: {user_id}"
                        )
                    except Exception as notify_error:
                        logger.warning(f"Не удалось уведомить администратора {admin['telegram_id']}: {notify_error}")
            
            await message.answer("Запрос отправлен. Ожидайте подтверждения.",
                               reply_markup=self.get_main_keyboard())
            await state.clear()

    async def process_employee_admin_data(self, message: types.Message, state: FSMContext):
        """Создание или обновление сотрудника администратором"""
        state_data = await state.get_data()
        action = state_data.get("employee_action", "add")
        target_employee_id = state_data.get("employee_id")

        parts = [part.strip() for part in message.text.split("|")]
        full_name = parts[0] if len(parts) > 0 and parts[0] else None
        position = parts[1] if len(parts) > 1 and parts[1] else None
        telegram_id = int(parts[2]) if len(parts) > 2 and parts[2] else None
        is_admin = self._parse_bool(parts[3], False) if len(parts) > 3 else False
        is_active = self._parse_bool(parts[4], True) if len(parts) > 4 else True
        is_approved = self._parse_bool(parts[5], True) if len(parts) > 5 else True

        if action == "add" and (not full_name or not telegram_id):
            await message.answer("Необходимо указать ФИО и Telegram ID")
            return

        async with self.pool.acquire() as conn:
            if action == "edit" and target_employee_id:
                current = await conn.fetchrow('''
                    SELECT full_name, position, telegram_id, is_admin, is_active, is_approved
                    FROM employees WHERE id = $1
                ''', target_employee_id)
                if not current:
                    await message.answer("Сотрудник не найден", reply_markup=self.get_main_keyboard())
                    await state.clear()
                    return

                new_full_name = full_name or current['full_name']
                new_position = position if position is not None else current['position']
                new_telegram_id = telegram_id or current['telegram_id']
                new_is_admin = is_admin if len(parts) > 3 else current['is_admin']
                new_is_active = is_active if len(parts) > 4 else current['is_active']
                new_is_approved = is_approved if len(parts) > 5 else current['is_approved']

                await conn.execute('''
                    UPDATE employees
                    SET full_name = $1, position = $2, telegram_id = $3,
                        is_admin = $4, is_active = $5, is_approved = $6
                    WHERE id = $7
                    ''', new_full_name, new_position, new_telegram_id, new_is_admin, new_is_active, new_is_approved, target_employee_id)

                updated = await conn.fetchrow(
                    '''
                    SELECT id, full_name, position, telegram_id, is_admin, is_active, is_approved
                    FROM employees WHERE id = $1
                    ''',
                    target_employee_id,
                )
                summary = self._format_employee_summary(updated) if updated else "Данные сотрудника обновлены"
                await message.answer(
                    f"Данные сотрудника обновлены:\n{summary}",
                    reply_markup=await self.build_main_keyboard_for_user(message.from_user.id),
                )
            else:
                await conn.execute('''
                    INSERT INTO employees (full_name, position, telegram_id, is_admin, is_active, is_approved)
                    VALUES ($1, $2, $3, $4, $5, $6)
                    ON CONFLICT (telegram_id) DO UPDATE SET
                        full_name = EXCLUDED.full_name,
                        position = EXCLUDED.position,
                        is_admin = EXCLUDED.is_admin,
                        is_active = EXCLUDED.is_active,
                        is_approved = EXCLUDED.is_approved
                ''', full_name, position, telegram_id, is_admin, is_active, is_approved)

                saved = await conn.fetchrow(
                    '''
                    SELECT id, full_name, position, telegram_id, is_admin, is_active, is_approved
                    FROM employees WHERE telegram_id = $1
                    ''',
                    telegram_id,
                )
                summary = self._format_employee_summary(saved) if saved else "Сотрудник сохранен"
                await message.answer(f"Сотрудник сохранен:\n{summary}", reply_markup=self.get_main_keyboard())

        await state.clear()
    
    async def process_sick_reason(self, message: types.Message, state: FSMContext):
        """Обработка причины больничного"""
        data = await state.get_data()
        user_id = data.get("user_id")
        reason = message.text

        today = today_moscow()
        async with self.pool.acquire() as conn:
            working_today = await conn.fetchval(
                '''
                SELECT 1 FROM time_logs
                WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1)
                  AND date = $2 AND check_in IS NOT NULL
                LIMIT 1
                ''',
                user_id,
                today,
            )
            if working_today:
                await message.answer(
                    "Приход уже отмечен, больничный недоступен.",
                    reply_markup=await self.build_main_keyboard_for_user(user_id),
                )
                await state.clear()
                return
            await conn.execute('''
                INSERT INTO time_logs (employee_id, date, status, notes)
                VALUES ((SELECT id FROM employees WHERE telegram_id = $1), $2, 'sick', $3)
            ''', user_id, today, reason)
        
        await message.answer(
            f"Больничный отмечен\nПричина: {reason if reason else 'не указана'}",
            reply_markup=await self.build_main_keyboard_for_user(user_id),
        )
        await state.clear()

    async def start_object_input(self, callback: types.CallbackQuery, state: FSMContext, action: str = "add"):
        """Подготовка ввода данных объекта"""
        object_id = None
        if action == "edit" and callback.data.startswith("admin_obj_edit_"):
            object_id = int(callback.data.rsplit("_", 1)[1])

        await state.set_state(Form.waiting_for_object_data)
        await state.update_data(object_action=action, object_id=object_id)
        await callback.message.answer(
            "Введите данные объекта в формате:\n"
            "Название | Адрес | Широта | Долгота | Радиус (м)\n"
            "Координаты и радиус можно не указывать",
            reply_markup=types.ReplyKeyboardRemove(),
        )

    async def choose_object_for_action(self, callback: types.CallbackQuery, action: str):
        """Выбор объекта для редактирования/удаления"""
        async with self.pool.acquire() as conn:
            objects = await conn.fetch('SELECT id, name FROM objects ORDER BY name')

        if not objects:
            await callback.message.answer("Список объектов пуст")
            return

        keyboard = InlineKeyboardBuilder()
        for obj in objects:
            suffix = f"admin_obj_{action}_{obj['id']}"
            keyboard.button(text=obj['name'], callback_data=suffix)
        keyboard.adjust(1)
        await callback.message.answer("Выберите объект:", reply_markup=keyboard.as_markup())

    async def delete_object(self, callback: types.CallbackQuery):
        """Удаление объекта"""
        try:
            obj_id = int(callback.data.rsplit("_", 1)[1])
        except Exception:
            await callback.message.answer("Не удалось определить объект")
            return

        async with self.pool.acquire() as conn:
            await conn.execute('UPDATE time_logs SET object_id = NULL WHERE object_id = $1', obj_id)
            deleted = await conn.execute('DELETE FROM objects WHERE id = $1', obj_id)

        await callback.message.answer(f"Объект удален ({deleted})")

    def _parse_bool(self, value: Optional[str], default: bool = False) -> bool:
        if value is None:
            return default
        return value.strip().lower() in {"1", "true", "yes", "да", "y", "on"}

    def _format_employee_summary(self, emp: asyncpg.Record) -> str:
        """Форматирует карточку сотрудника для быстрого просмотра"""
        role = "Администратор" if emp['is_admin'] else "Сотрудник"
        status_parts = []
        status_parts.append("Активен" if emp['is_active'] else "Деактивирован")
        status_parts.append("Одобрен" if emp['is_approved'] else "Не одобрен")
        position = emp['position'] or "Не указана"
        return (
            f"{emp['full_name']}\n"
            f"Должность: {position}\n"
            f"Роль: {role}\n"
            f"Telegram ID: {emp['telegram_id']}\n"
            f"Статус: {', '.join(status_parts)}"
        )

    async def start_employee_input(self, callback: types.CallbackQuery, state: FSMContext, action: str = "add"):
        """Подготовка ввода данных сотрудника"""
        employee_id = None
        if action == "edit" and callback.data.startswith("admin_emp_edit_"):
            employee_id = int(callback.data.rsplit("_", 1)[1])

        current_text = ""
        if employee_id:
            async with self.pool.acquire() as conn:
                current = await conn.fetchrow(
                    '''
                    SELECT id, full_name, position, telegram_id, is_admin, is_active, is_approved
                    FROM employees WHERE id = $1
                    ''',
                    employee_id,
                )
            if current:
                current_text = (
                    "\n\nТекущие данные:\n" +
                    self._format_employee_summary(current) +
                    "\n\nОставьте поле пустым между символами | чтобы не менять значение."
                )

        await state.set_state(Form.waiting_for_employee_admin_data)
        await state.update_data(employee_action=action, employee_id=employee_id)
        await callback.message.answer(
            "Введите данные сотрудника в формате:\n"
            "ФИО | Должность | Telegram ID | Админ (0/1) | Активен (0/1) | Одобрен (0/1)\n"
            "Поля admin/active/approved можно пропустить, по умолчанию: 0/1/1" + current_text,
            reply_markup=types.ReplyKeyboardRemove(),
        )

    async def choose_employee_for_action(self, callback: types.CallbackQuery, action: str):
        async with self.pool.acquire() as conn:
            employees = await conn.fetch('SELECT id, full_name FROM employees ORDER BY full_name')

        if not employees:
            await callback.message.answer("Список сотрудников пуст")
            return

        keyboard = InlineKeyboardBuilder()
        for emp in employees:
            suffix = f"admin_emp_view_{emp['id']}" if action == "edit" else f"admin_emp_{action}_{emp['id']}"
            keyboard.button(text=emp['full_name'], callback_data=suffix)
        keyboard.adjust(1)
        await callback.message.answer("Выберите сотрудника:", reply_markup=keyboard.as_markup())

    async def show_employee_card(self, callback: types.CallbackQuery):
        """Карточка сотрудника с быстрыми действиями"""
        try:
            emp_id = int(callback.data.rsplit("_", 1)[1])
        except Exception:
            await callback.message.answer("Не удалось определить сотрудника")
            return

        async with self.pool.acquire() as conn:
            emp = await conn.fetchrow(
                '''
                SELECT id, full_name, position, telegram_id, is_admin, is_active, is_approved
                FROM employees WHERE id = $1
                ''',
                emp_id,
            )

        if not emp:
            await callback.message.answer("Сотрудник не найден")
            return

        builder = InlineKeyboardBuilder()
        builder.button(text="Редактировать поля", callback_data=f"admin_emp_edit_{emp_id}")
        builder.button(
            text="Переключить роль",
            callback_data=f"admin_emp_toggle_admin_{emp_id}",
        )
        builder.button(
            text="Активен/Стоп",
            callback_data=f"admin_emp_toggle_active_{emp_id}",
        )
        builder.button(
            text="Одобрить/Скрыть",
            callback_data=f"admin_emp_toggle_approved_{emp_id}",
        )
        builder.button(text="Удалить", callback_data=f"admin_emp_delete_{emp_id}")
        builder.button(text="Назад к списку", callback_data="admin_edit_employee")
        builder.adjust(2)

        await callback.message.answer(
            self._format_employee_summary(emp),
            reply_markup=builder.as_markup(),
        )

    async def toggle_employee_flag(self, callback: types.CallbackQuery):
        """Быстрое переключение флагов сотрудника"""
        try:
            _, _, _, field, emp_id_str = callback.data.split("_", maxsplit=4)
            emp_id = int(emp_id_str)
        except Exception:
            await callback.answer("Не удалось определить действие", show_alert=True)
            return

        field_map = {
            "admin": "is_admin",
            "active": "is_active",
            "approved": "is_approved",
        }
        column = field_map.get(field)
        if not column:
            await callback.answer("Неизвестный флаг", show_alert=True)
            return

        async with self.pool.acquire() as conn:
            emp = await conn.fetchrow(
                f'''
                SELECT id, telegram_id, {column} FROM employees WHERE id = $1
                ''',
                emp_id,
            )
            if not emp:
                await callback.answer("Сотрудник не найден", show_alert=True)
                return

            new_value = not emp[column]
            if emp['telegram_id'] in Config.ADMIN_IDS and not new_value:
                await callback.answer("Нельзя отключить главного администратора", show_alert=True)
                return

            await conn.execute(
                f'''UPDATE employees SET {column} = $1 WHERE id = $2''',
                new_value,
                emp_id,
            )

        await self.show_employee_card(callback)

    async def delete_employee(self, callback: types.CallbackQuery):
        try:
            emp_id = int(callback.data.rsplit("_", 1)[1])
        except Exception:
            await callback.message.answer("Не удалось определить сотрудника")
            return

        async with self.pool.acquire() as conn:
            employee = await conn.fetchrow('SELECT telegram_id FROM employees WHERE id = $1', emp_id)
            if employee and employee['telegram_id'] in Config.ADMIN_IDS:
                await callback.message.answer("Нельзя удалить главного администратора")
                return

            await conn.execute(
                '''
                UPDATE employees
                SET is_active = FALSE, is_approved = FALSE
                WHERE id = $1
                ''',
                emp_id,
            )

        await callback.message.answer("Сотрудник деактивирован")
    
    async def process_object_data(self, message: types.Message, state: FSMContext):
        """Обработка данных объекта"""
        state_data = await state.get_data()
        action = state_data.get("object_action", "add")
        target_object_id = state_data.get("object_id")
        parts = message.text.strip().split("|")
        name = parts[0].strip()
        address = parts[1].strip() if len(parts) > 1 else None
        
        lat = lon = radius = None
        if len(parts) > 2: 
            lat = float(parts[2]) if parts[2].strip() else None
        if len(parts) > 3:
            lon = float(parts[3]) if parts[3].strip() else None
        if len(parts) > 4:
            radius = int(parts[4]) if parts[4].strip() else 500

        async with self.pool.acquire() as conn:
            if action == "edit" and target_object_id:
                current = await conn.fetchrow('SELECT name, address, latitude, longitude, radius FROM objects WHERE id = $1', target_object_id)
                if not current:
                    await message.answer("Объект не найден", reply_markup=self.get_main_keyboard())
                    await state.clear()
                    return

                updated_name = name or current['name']
                updated_address = address if address is not None else current['address']
                updated_lat = lat if lat is not None else current['latitude']
                updated_lon = lon if lon is not None else current['longitude']
                updated_radius = radius if radius is not None else current['radius'] or 500

                await conn.execute('''
                    UPDATE objects
                    SET name = $1, address = $2, latitude = $3, longitude = $4, radius = $5
                    WHERE id = $6
                ''', updated_name, updated_address, updated_lat, updated_lon, updated_radius, target_object_id)

                await message.answer(
                    f"Объект обновлен: {updated_name}",
                    reply_markup=self.get_main_keyboard(),
                )
            else:
                await conn.execute('''
                    INSERT INTO objects (name, address, latitude, longitude, radius)
                    VALUES ($1, $2, $3, $4, $5)
                ''', name, address, lat, lon, radius)

                await message.answer(f"Объект добавлен: {name}",
                                   reply_markup=self.get_main_keyboard())
        await state.clear()
    
    # Обработчики геолокации и колбэков
    async def handle_location(self, message: types.Message, state: FSMContext):
        """Обработка геолокации"""
        data = await state.get_data()
        action = data.get("action")
        selected_object_id = data.get("selected_object_id")
        location = message.location
        message_time_msk = self.normalize_message_time(message.date)

        if not action:
            await message.answer("Не удалось определить действие. Попробуйте снова через меню.",
                               reply_markup=self.get_main_keyboard())
            await state.clear()
            return

        async with self.pool.acquire() as conn:
            target_object = None
            if selected_object_id:
                target_object = await conn.fetchrow(
                    'SELECT id, name, latitude, longitude, radius FROM objects WHERE id = $1',
                    selected_object_id,
                )
            else:
                objects = await conn.fetch('SELECT id, name, latitude, longitude, radius FROM objects')
                nearest = None
                min_dist = float('inf')
                for obj in objects:
                    if obj['latitude'] and obj['longitude']:
                        dist = calculate_distance(
                            location.latitude, location.longitude,
                            obj['latitude'], obj['longitude']
                        )
                        if dist < min_dist:
                            min_dist = dist
                            nearest = obj
                target_object = nearest

        if not target_object:
            await self.show_objects(message, state, action)
            await state.update_data(lat=location.latitude, lon=location.longitude)
            return

        distance = None
        if target_object['latitude'] and target_object['longitude']:
            distance = calculate_distance(
                location.latitude,
                location.longitude,
                target_object['latitude'],
                target_object['longitude'],
            )

        accuracy = location.horizontal_accuracy
        suspicious_reasons = []
        age_seconds = (self.moscow_now() - message_time_msk).total_seconds()
        if age_seconds > Config.LOCATION_MAX_AGE_SECONDS:
            suspicious_reasons.append(
                f"метка старше {Config.LOCATION_MAX_AGE_SECONDS} c (получена {message_time_msk.strftime('%H:%M:%S')})"
            )
        if accuracy and accuracy > max(Config.LOCATION_RADIUS, Config.LOCATION_MAX_ACCURACY):
            suspicious_reasons.append(
                f"точность {accuracy:.0f} м превышает допустимые {max(Config.LOCATION_RADIUS, Config.LOCATION_MAX_ACCURACY):.0f} м"
            )

        if suspicious_reasons:
            reason_text = "; ".join(suspicious_reasons)
            await self.log_location_attempt(
                user_id=message.from_user.id,
                object_id=target_object['id'] if target_object else None,
                action=action,
                location=location,
                distance=distance,
                accuracy=accuracy,
                message_time=message_time_msk,
                is_suspicious=True,
                reason=reason_text,
            )
            await message.answer(
                "Геолокация выглядит недостоверной. Отправьте актуальную точку через кнопку "
                "«Отправить геолокацию» и убедитесь, что GPS включен.",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[[KeyboardButton(text="Отправить геолокацию", request_location=True)],
                              [KeyboardButton(text="Отмена")]],
                    resize_keyboard=True,
                ),
            )
            await state.update_data(selected_object_id=target_object['id'])
            await state.set_state(Form.waiting_for_location)
            return

        if (
            distance is not None
            and distance != float('inf')
            and distance > Config.LOCATION_RADIUS
        ):
            await message.answer(
                f"Вы находитесь дальше {Config.LOCATION_RADIUS} м от выбранного объекта. Подойдите ближе и отправьте геолокацию еще раз.",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[
                        [KeyboardButton(text="Отправить геолокацию", request_location=True)],
                        [KeyboardButton(text="Не получилось отметить — я на объекте")],
                        [KeyboardButton(text="Отмена")],
                    ],
                    resize_keyboard=True,
                ),
            )
            await state.update_data(selected_object_id=target_object['id'])
            await state.set_state(Form.waiting_for_location)
            return

        if (
            distance is not None
            and distance != float('inf')
            and distance > Config.LOCATION_RADIUS
        ):
            await state.update_data(
                manual_context={
                    "action": action,
                    "object_id": target_object['id'],
                    "log_id": data.get("log_id"),
                    "lat": location.latitude,
                    "lon": location.longitude,
                    "distance": distance,
                }
            )
            await message.answer(
                f"Вы находитесь дальше {Config.LOCATION_RADIUS} м от выбранного объекта. Подойдите ближе и отправьте геолокацию еще раз.",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[
                        [KeyboardButton(text="Отправить геолокацию", request_location=True)],
                        [KeyboardButton(text="Не получилось отметить — я на объекте")],
                        [KeyboardButton(text="Отмена")],
                    ],
                    resize_keyboard=True,
                ),
            )
            await self.log_location_attempt(
                user_id=message.from_user.id,
                object_id=target_object['id'] if target_object else None,
                action=action,
                location=location,
                distance=distance,
                accuracy=accuracy,
                message_time=message_time_msk,
                is_suspicious=True,
                reason=f"дистанция {distance:.0f} м превышает допуск {Config.LOCATION_RADIUS} м",
            )
            await state.update_data(selected_object_id=target_object['id'])
            await state.set_state(Form.waiting_for_location)
            return

        await self.log_location_attempt(
            user_id=message.from_user.id,
            object_id=target_object['id'] if target_object else None,
            action=action,
            location=location,
            distance=distance,
            accuracy=accuracy,
            message_time=message_time_msk,
            is_suspicious=False,
            reason=None,
        )

        if action == "checkin":
            await self.process_checkin_with_location(message.from_user.id, location, target_object, distance or 0)
            await state.clear()
        elif action == "checkout":
            await self.process_checkout_with_location(data.get("log_id"), location, target_object, distance or 0)
            await state.clear()

    async def log_location_attempt(
        self,
        user_id: int,
        object_id: Optional[int],
        action: str,
        location: types.Location,
        distance: Optional[float],
        accuracy: Optional[float],
        message_time: datetime,
        is_suspicious: bool,
        reason: Optional[str],
    ):
        """Фиксирует попытку отметки с геолокацией для антифрода и аудита"""
        try:
            async with self.pool.acquire() as conn:
                await conn.execute(
                    '''
                    INSERT INTO location_audit (
                        employee_id, object_id, action, latitude, longitude,
                        distance, accuracy, message_time, is_suspicious, reason
                    )
                    VALUES (
                        (SELECT id FROM employees WHERE telegram_id = $1),
                        $2, $3, $4, $5, $6, $7, $8, $9, $10
                    )
                    ''',
                    user_id,
                    object_id,
                    action,
                    location.latitude,
                    location.longitude,
                    int(distance or 0) if distance not in (None, float('inf')) else None,
                    float(accuracy) if accuracy is not None else None,
                    message_time.replace(tzinfo=None),
                    is_suspicious,
                    reason,
                )
        except Exception as log_err:
            logger.error(f"Не удалось зафиксировать попытку геолокации: {log_err}")
    
    async def _apply_manual_time(
        self,
        user_id: int,
        object_id: Optional[int],
        action: str,
        log_id: Optional[int],
        note: str,
        photo_id: str,
        requested_at: datetime,
        lat: Optional[float],
        lon: Optional[float],
        distance: Optional[float],
        status_label: str = "manual_confirmed",
    ) -> Optional[int]:
        """Применяет отметку вручную (авто или после одобрения)"""
        extra_note = f"Ручное подтверждение: {note}".strip()
        if distance not in (None, float("inf")):
            extra_note += f" (расстояние {distance:.0f} м)"
        if photo_id:
            extra_note += f"\nФото: {photo_id}"

        if action == "checkin":
            async with self.pool.acquire() as conn:
                await conn.execute(
                    '''
                    INSERT INTO time_logs (
                        employee_id, object_id, date, check_in, check_in_lat, check_in_lon, status, notes
                    )
                    VALUES ((SELECT id FROM employees WHERE telegram_id = $1), $2, $3, $4, $5, $6, $7, $8)
                    ON CONFLICT DO NOTHING
                    ''',
                    user_id,
                    object_id,
                    requested_at.date(),
                    requested_at,
                    lat,
                    lon,
                    status_label,
                    extra_note,
                )
                new_log_id = await conn.fetchval(
                    '''
                    SELECT id FROM time_logs
                    WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1)
                      AND date = $2
                    ORDER BY id DESC
                    LIMIT 1
                    ''',
                    user_id,
                    requested_at.date(),
                )
                return new_log_id

        elif action == "checkout":
            async with self.pool.acquire() as conn:
                if not log_id:
                    log_id = await conn.fetchval(
                        '''
                        SELECT id FROM time_logs
                        WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1)
                          AND date = $2 AND check_out IS NULL
                        ORDER BY id DESC LIMIT 1
                        ''',
                        user_id,
                        requested_at.date(),
                    )
                if not log_id:
                    return None

                check_in = await conn.fetchval('SELECT check_in FROM time_logs WHERE id = $1', log_id)
                hours = 0
                if check_in:
                    hours = (requested_at - check_in).total_seconds() / 3600
                hours = min(max(0, math.ceil(hours)), Config.MAX_WORK_HOURS)

                await conn.execute(
                    '''
                    UPDATE time_logs
                    SET check_out = $1,
                        check_out_lat = $2,
                        check_out_lon = $3,
                        hours_worked = $4,
                        status = $5,
                        notes = COALESCE(notes, '') || '\n' || $6
                    WHERE id = $7
                    ''',
                    requested_at,
                    lat,
                    lon,
                    hours,
                    status_label,
                    extra_note,
                    log_id,
                )
                return log_id
        return None

    async def _save_manual_request(
        self,
        user_id: int,
        object_id: Optional[int],
        action: str,
        status: str,
        note: str,
        photo_id: str,
        lat: Optional[float],
        lon: Optional[float],
        distance: Optional[float],
        log_id: Optional[int],
    ) -> int:
        async with self.pool.acquire() as conn:
            req_id = await conn.fetchval(
                '''
                INSERT INTO manual_presence_requests (
                    employee_id, object_id, action, status,
                    note, photo_file_id, latitude, longitude, distance, log_id
                )
                VALUES (
                    (SELECT id FROM employees WHERE telegram_id = $1),
                    $2, $3, $4, $5, $6, $7, $8, $9, $10
                )
                RETURNING id
                ''',
                user_id,
                object_id,
                action,
                status,
                note,
                photo_id,
                lat,
                lon,
                int(distance) if distance not in (None, float("inf")) else None,
                log_id,
            )
            return req_id

    async def notify_admins_manual_request(self, req_id: int):
        """Уведомление администраторов о новой ручной отметке"""
        async with self.pool.acquire() as conn:
            req = await conn.fetchrow(
                '''
                SELECT m.id, m.action, m.note, m.photo_file_id, m.distance, m.latitude, m.longitude,
                       e.full_name, o.name as object_name
                FROM manual_presence_requests m
                JOIN employees e ON m.employee_id = e.id
                LEFT JOIN objects o ON m.object_id = o.id
                WHERE m.id = $1
                ''',
                req_id,
            )
        if not req:
            return

        text = (
            f"Ручное подтверждение #{req['id']}\n"
            f"Сотрудник: {req['full_name']}\n"
            f"Действие: {req['action']}\n"
            f"Объект: {req['object_name'] or '—'}\n"
            f"Дистанция: {req['distance'] or '—'} м\n"
            f"Комментарий: {req['note'] or '—'}"
        )
        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(text="Одобрить", callback_data=f"admin_manual_approve_{req['id']}"),
                    InlineKeyboardButton(text="Отклонить", callback_data=f"admin_manual_reject_{req['id']}"),
                ]
            ]
        )
        for admin_id in Config.ADMIN_IDS:
            try:
                if req['photo_file_id']:
                    await self.bot.send_photo(
                        admin_id,
                        req['photo_file_id'],
                        caption=text,
                        reply_markup=keyboard,
                    )
                else:
                    await self.bot.send_message(admin_id, text, reply_markup=keyboard)
            except Exception as err:
                logger.error(f"Не удалось отправить запрос {req_id} администратору {admin_id}: {err}")

    async def show_manual_requests(self, callback: types.CallbackQuery):
        """Показать ожидающие ручные подтверждения"""
        async with self.pool.acquire() as conn:
            rows = await conn.fetch(
                '''
                SELECT m.id, e.full_name, m.action, m.note, m.distance, m.photo_file_id, o.name as object_name
                FROM manual_presence_requests m
                JOIN employees e ON m.employee_id = e.id
                LEFT JOIN objects o ON m.object_id = o.id
                WHERE m.status = 'pending'
                ORDER BY m.id DESC
                '''
            )

        if not rows:
            await callback.message.answer("Нет запросов на подтверждение присутствия.")
            return

        for row in rows:
            text = (
                f"Запрос #{row['id']}\n"
                f"Сотрудник: {row['full_name']}\n"
                f"Действие: {row['action']}\n"
                f"Объект: {row['object_name'] or '—'}\n"
                f"Дистанция: {row['distance'] or '—'} м\n"
                f"Комментарий: {row['note'] or '—'}"
            )
            keyboard = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(text="Одобрить", callback_data=f"admin_manual_approve_{row['id']}"),
                        InlineKeyboardButton(text="Отклонить", callback_data=f"admin_manual_reject_{row['id']}"),
                    ]
                ]
            )
            if row['photo_file_id']:
                await callback.message.answer_photo(row['photo_file_id'], caption=text, reply_markup=keyboard)
            else:
                await callback.message.answer(text, reply_markup=keyboard)

    async def approve_manual_request(self, callback: types.CallbackQuery):
        """Одобрить ручную отметку"""
        try:
            req_id = int(callback.data.rsplit("_", 1)[1])
        except Exception:
            await callback.answer("Некорректный запрос", show_alert=True)
            return

        async with self.pool.acquire() as conn:
            req = await conn.fetchrow(
                '''
                SELECT m.*, e.telegram_id, e.full_name
                FROM manual_presence_requests m
                JOIN employees e ON m.employee_id = e.id
                WHERE m.id = $1
                ''',
                req_id,
            )
        if not req:
            await callback.answer("Запрос не найден", show_alert=True)
            return
        if req['status'] != 'pending':
            await callback.answer("Уже обработано", show_alert=True)
            return

        log_id = await self._apply_manual_time(
            user_id=req['telegram_id'],
            object_id=req['object_id'],
            action=req['action'],
            log_id=req['log_id'],
            note=req['note'] or "",
            photo_id=req['photo_file_id'] or "",
            requested_at=req['requested_at'] or self.moscow_now_naive(),
            lat=float(req['latitude']) if req['latitude'] is not None else None,
            lon=float(req['longitude']) if req['longitude'] is not None else None,
            distance=float(req['distance']) if req['distance'] is not None else None,
            status_label="manual_admin",
        )

        async with self.pool.acquire() as conn:
            await conn.execute(
                '''
                UPDATE manual_presence_requests
                SET status = 'approved', admin_id = $1, processed_at = $2, log_id = COALESCE(log_id, $3)
                WHERE id = $4
                ''',
                callback.from_user.id,
                self.moscow_now_naive(),
                log_id,
                req_id,
            )

        await callback.answer("Одобрено")
        try:
            await self.bot.send_message(
                req['telegram_id'],
                "Ручная отметка одобрена администратором.",
                reply_markup=await self.build_main_keyboard_for_user(req['telegram_id']),
            )
        except Exception as err:
            logger.error(f"Не удалось уведомить пользователя о одобрении {req_id}: {err}")

    async def reject_manual_request(self, callback: types.CallbackQuery):
        """Отклонить ручную отметку"""
        try:
            req_id = int(callback.data.rsplit("_", 1)[1])
        except Exception:
            await callback.answer("Некорректный запрос", show_alert=True)
            return

        async with self.pool.acquire() as conn:
            req = await conn.fetchrow(
                '''
                SELECT m.id, m.status, e.telegram_id
                FROM manual_presence_requests m
                JOIN employees e ON m.employee_id = e.id
                WHERE m.id = $1
                ''',
                req_id,
            )
        if not req:
            await callback.answer("Запрос не найден", show_alert=True)
            return
        if req['status'] != 'pending':
            await callback.answer("Уже обработано", show_alert=True)
            return

        async with self.pool.acquire() as conn:
            await conn.execute(
                '''
                UPDATE manual_presence_requests
                SET status = 'rejected', admin_id = $1, processed_at = $2
                WHERE id = $3
                ''',
                callback.from_user.id,
                self.moscow_now_naive(),
                req_id,
            )

        await callback.answer("Отклонено")
        try:
            await self.bot.send_message(
                req['telegram_id'],
                "Ручная отметка отклонена администратором.",
                reply_markup=await self.build_main_keyboard_for_user(req['telegram_id']),
            )
        except Exception as err:
            logger.error(f"Не удалось уведомить пользователя об отклонении {req_id}: {err}")

    # Методы обработки данных
    async def process_checkin_manual(
        self,
        user_id: int,
        obj: asyncpg.Record,
        lat: Optional[float] = None,
        lon: Optional[float] = None,
        distance: Optional[float] = None,
    ):
        """Обработка прихода с выбранным объектом"""
        now = self.moscow_now_naive()
        status = 'work'
        work_start = datetime.combine(now.date(), time(Config.WORK_START_HOUR, 0))
        if now > work_start + timedelta(minutes=15):
            status = 'late'

        async with self.pool.acquire() as conn:
            await conn.execute(
                '''
                INSERT INTO time_logs (employee_id, object_id, date, check_in, check_in_lat, check_in_lon, status)
                VALUES ((SELECT id FROM employees WHERE telegram_id = $1), $2, $3, $4, $5, $6, $7)
                ''',
                user_id,
                obj['id'],
                now.date(),
                now,
                lat,
                lon,
                status,
            )
            is_admin = await conn.fetchval('SELECT is_admin FROM employees WHERE telegram_id = $1', user_id)

        keyboard = self.get_main_keyboard(is_admin=is_admin)

        message_text = f"Приход отмечен!\nВремя: {now.strftime('%H:%M')}\nОбъект: {obj['name']}"
        if distance is not None:
            message_text += f"\nРасстояние: {distance:.0f} м"
        if status == 'late':
            message_text += "\nВы опоздали!"

        await self.bot.send_message(user_id, message_text, reply_markup=keyboard)
        logger.info(
            "Отметка прихода: %s (%s) на объекте %s в %s, расстояние %.0f м",
            user_id,
            obj['id'],
            obj['name'],
            now.strftime('%H:%M'),
            distance or 0,
        )

    async def process_checkin_with_location(self, user_id: int, location, obj, distance: float):
        """Обработка прихода с геолокацией"""
        await self.process_checkin_manual(user_id, obj, location.latitude, location.longitude, distance)
    
    async def process_checkout_manual(
        self,
        log_id: int,
        obj: asyncpg.Record,
        user_id: int,
        lat: Optional[float] = None,
        lon: Optional[float] = None,
    ):
        """Обработка ухода с выбранным объектом"""
        now = self.moscow_now_naive()

        async with self.pool.acquire() as conn:
            check_in = await conn.fetchval('SELECT check_in FROM time_logs WHERE id = $1', log_id)
            hours = 0
            if check_in:
                hours = (now - check_in).total_seconds() / 3600
            hours = min(max(0, math.ceil(hours)), Config.MAX_WORK_HOURS)

            await conn.execute(
                '''
                UPDATE time_logs SET check_out = $1, check_out_lat = $2, check_out_lon = $3,
                hours_worked = $4, object_id = $5 WHERE id = $6
                ''',
                now,
                lat,
                lon,
                hours,
                obj['id'],
                log_id,
            )

            is_admin = await conn.fetchval('SELECT is_admin FROM employees WHERE telegram_id = $1', user_id)

        keyboard = await self.build_main_keyboard_for_user(user_id)
        await self.bot.send_message(
            user_id,
            f"Уход отмечен!\nВремя: {now.strftime('%H:%M')}\nОбъект: {obj['name']}\nОтработано: {hours} ч.",
            reply_markup=keyboard,
        )
        logger.info(
            "Отметка ухода: %s (%s) на объекте %s в %s, часов: %.0f",
            user_id,
            obj['id'],
            obj['name'],
            now.strftime('%H:%M'),
            hours,
        )

    async def process_checkout_with_location(self, log_id: int, location, obj, distance: float):
        """Обработка ухода с геолокацией"""
        if not log_id:
            logger.warning("Не удалось определить запись времени для отметки ухода")
            return
        async with self.pool.acquire() as conn:
            user_id = await conn.fetchval(
                '''
                SELECT telegram_id FROM employees WHERE id = (
                    SELECT employee_id FROM time_logs WHERE id = $1
                )
                ''',
                log_id,
            )
        if not user_id:
            logger.warning("Не удалось определить пользователя для отметки ухода")
            return
        await self.process_checkout_manual(log_id, obj, user_id, location.latitude, location.longitude)
    
    async def process_checkout_simple(self, user_id: int, log_id: int):
        """Уход без геолокации"""
        now = self.moscow_now_naive()
        
        async with self.pool.acquire() as conn:
            # Получаем время прихода
            check_in = await conn.fetchval('SELECT check_in FROM time_logs WHERE id = $1', log_id)
            
            if check_in:
                # Исправленный расчет часов
                hours = (now - check_in).total_seconds() / 3600
            else:
                hours = 0

            hours = min(max(0, math.ceil(hours)), Config.MAX_WORK_HOURS)
            
            # Обновляем запись
            await conn.execute('''
                UPDATE time_logs 
                SET check_out = $1, hours_worked = $2 
                WHERE id = $3
            ''', now, hours, log_id)
            
            # Получаем статус администратора
        keyboard = await self.build_main_keyboard_for_user(user_id)
        await self.bot.send_message(
            user_id,
            f"Уход отмечен в {now.strftime('%H:%M')}\nОтработано: {hours} ч.",
            reply_markup=keyboard
        )
        logger.info(
            "Уход без геолокации: %s запись %s в %s, часов: %.0f",
            user_id,
            log_id,
            now.strftime('%H:%M'),
            hours,
        )
    
    # Исправляем метод одобрения запроса
    async def handle_approval(self, callback: types.CallbackQuery):
        """Одобрить запрос"""
        req_id = int(callback.data.split("_")[1])

        async with self.pool.acquire() as conn:
            await self.ensure_position_columns(conn)
            try:
                req = await conn.fetchrow('''
                    SELECT telegram_id, full_name, position FROM access_requests WHERE id = $1
                ''', req_id)
            except asyncpg.UndefinedColumnError:
                await self.ensure_position_columns(conn)
                req = await conn.fetchrow('''
                    SELECT telegram_id, full_name, position FROM access_requests WHERE id = $1
                ''', req_id)
            
            if not req:
                await callback.answer("Запрос не найден")
                return
            
            # Обновляем статус запроса
            await conn.execute('''
                UPDATE access_requests SET status = 'approved' WHERE id = $1
            ''', req_id)
            
            # Добавляем сотрудника с должностью
            await conn.execute('''
                INSERT INTO employees (telegram_id, full_name, position, is_approved)
                VALUES ($1, $2, $3, TRUE)
                ON CONFLICT (telegram_id) 
                DO UPDATE SET 
                    is_approved = TRUE, 
                    full_name = EXCLUDED.full_name,
                    position = EXCLUDED.position
            ''', req['telegram_id'], req['full_name'], req['position'])
            
            is_admin = await conn.fetchval('SELECT is_admin FROM employees WHERE telegram_id = $1', 
                                          callback.from_user.id)
        
        await callback.message.edit_text(f"Запрос одобрен\nФИО: {req['full_name']}\nДолжность: {req['position']}")
        
        try:
            keyboard = self.get_main_keyboard()
            await self.bot.send_message(
                req['telegram_id'], 
                "Ваш запрос на доступ одобрен! Используйте /start для начала работы.",
                reply_markup=keyboard
            )
        except Exception as e:
            logger.error(f"Ошибка отправки сообщения: {e}")
            
    async def handle_request_access(self, message: types.Message, state: FSMContext):
        """Запрос доступа"""
        user_id = message.from_user.id

        async with self.pool.acquire() as conn:
            await self.ensure_position_columns(conn)
            try:
                existing = await conn.fetchval('''
                    SELECT status FROM access_requests WHERE telegram_id = $1 ORDER BY id DESC LIMIT 1
                ''', user_id)
            except asyncpg.UndefinedColumnError:
                await self.ensure_position_columns(conn)
                existing = await conn.fetchval('''
                    SELECT status FROM access_requests WHERE telegram_id = $1 ORDER BY id DESC LIMIT 1
                ''', user_id)
            
            if existing == 'pending':
                await message.answer("Ваш запрос уже на рассмотрении.")
                return
        
        await message.answer("Введите ваше ФИО и должность в формате:\nФИО | Должность\n\nПример:\nИванов Иван Иванович | Менеджер")
        await state.set_state(Form.waiting_for_employee_name)
        await state.update_data(action="request_access", user_id=user_id)
            
    # Исправляем метод отклонения запроса
    async def handle_rejection(self, callback: types.CallbackQuery):
        """Отклонить запрос"""
        req_id = int(callback.data.split("_")[1])

        async with self.pool.acquire() as conn:
            await self.ensure_position_columns(conn)
            try:
                req = await conn.fetchrow('SELECT full_name FROM access_requests WHERE id = $1', req_id)
            except asyncpg.UndefinedColumnError:
                await self.ensure_position_columns(conn)
                req = await conn.fetchrow('SELECT full_name FROM access_requests WHERE id = $1', req_id)
            if req:
                await conn.execute("UPDATE access_requests SET status = 'rejected' WHERE id = $1", req_id)
        
        if req:
            await callback.message.edit_text(f"Запрос отклонен\nФИО: {req['full_name']}")
        else:
            await callback.message.edit_text("Запрос не найден")
    
    async def ask_timesheet_type(self, callback: types.CallbackQuery):
        """Предлагает выбор формата табеля"""
        builder = InlineKeyboardBuilder()
        builder.button(text="Общий табель", callback_data="admin_timesheet_overall")
        builder.button(text="Детализированный по объектам", callback_data="admin_timesheet_detailed")
        builder.adjust(1)
        await callback.message.answer(
            "Какой табель сформировать?",
            reply_markup=builder.as_markup(),
        )

    async def generate_timesheet(self, callback: types.CallbackQuery, include_objects: bool):
        """Сгенерировать табель указанного формата"""
        try:
            excel_file = await self._build_workbook(include_objects=include_objects)

            if excel_file:
                caption = "Общий табель без объектов" if not include_objects else "Детализированный табель"
                filename = "Общий_табель" if not include_objects else "Табель_по_объектам"
                await self.bot.send_document(
                    chat_id=callback.from_user.id,
                    document=types.BufferedInputFile(
                        excel_file.getvalue(),
                        filename=f"{filename}_{now_moscow().strftime('%Y_%m_%d')}.xlsx"
                    ),
                    caption=caption
                )
            else:
                await callback.message.answer("Ошибка формирования табеля")
        except Exception as e:
            logger.error(f"Ошибка: {e}")
            await callback.message.answer(f"Ошибка: {str(e)}")
    
    async def _build_workbook(self, include_objects: bool = True) -> Optional[io.BytesIO]:
        today = self.moscow_today()
        year, month = today.year, today.month
        days_in_month = monthrange(year, month)[1]
        month_label = date(year, month, 1).strftime('%B %Y')

        async with self.pool.acquire() as conn:
            employees = await conn.fetch(
                '''
                SELECT id, full_name, position, is_admin FROM employees
                WHERE is_approved = TRUE AND is_active = TRUE
                ORDER BY full_name
                '''
            )

            logs = await conn.fetch(
                '''
                SELECT tl.employee_id, tl.date, tl.check_in, tl.check_out,
                       tl.hours_worked, tl.status, tl.object_id,
                       o.name as object_name
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                LEFT JOIN objects o ON tl.object_id = o.id
                WHERE EXTRACT(YEAR FROM tl.date) = $1
                  AND EXTRACT(MONTH FROM tl.date) = $2
                ORDER BY e.full_name, tl.date
                ''',
                year,
                month,
            )

            objects = {row['id']: row for row in await conn.fetch('SELECT id, name FROM objects')}

        if not employees:
            return None

        logs_by_employee: Dict[int, Dict[int, Dict[str, float]]] = {}
        object_logs: Dict[int, Dict[int, Dict[int, Dict[str, float]]]] = {}

        for log in logs:
            emp_map = logs_by_employee.setdefault(log['employee_id'], {})
            day = log['date'].day
            if log['status'] == 'sick':
                emp_map[day] = {'status': 'sick'}
                continue

            hours = float(log['hours_worked'] or 0)
            existing = emp_map.get(day)
            if existing and existing.get('status') != 'sick':
                hours += existing.get('hours', 0)
            emp_map[day] = {'hours': hours}

            if log['object_id']:
                obj_emp = object_logs.setdefault(log['object_id'], {}).setdefault(log['employee_id'], {})
                prev = obj_emp.get(day, {}).get('hours', 0)
                obj_emp[day] = {'hours': prev + hours}

        wb = Workbook()

        ws_summary = wb.active
        ws_summary.title = "Общий табель"
        ws_summary.append([f"Общий табель за {month_label}"])
        ws_summary.append(["ФИО", "Должность"] + [str(d) for d in range(1, days_in_month + 1)] + ["Итого часов", "Дней"])

        for emp in employees:
            row = [emp['full_name'], emp['position'] or ""]
            total_hours = 0.0
            days_worked = 0
            emp_days = logs_by_employee.get(emp['id'], {})
            for day in range(1, days_in_month + 1):
                entry = emp_days.get(day)
                if not entry:
                    row.append("")
                    continue
                if entry.get('status') == 'sick':
                    row.append('Б')
                else:
                    hours_val = entry.get('hours', 0)
                    row.append(f"{hours_val:.1f}")
                    total_hours += hours_val
                    days_worked += 1
            row.append(total_hours)
            row.append(days_worked)
            ws_summary.append(row)

        if include_objects:
            for obj_id, emp_logs in object_logs.items():
                obj_name = objects.get(obj_id, {}).get('name', f"Объект {obj_id}")
                title = obj_name if len(obj_name) <= 25 else obj_name[:25] + f"_{obj_id}"
                ws = wb.create_sheet(title)
                ws.append([f"{obj_name} — {month_label}"])
                ws.append(["ФИО", "Должность"] + [str(d) for d in range(1, days_in_month + 1)] + ["Итого часов", "Дней"])

                for emp in employees:
                    emp_days = emp_logs.get(emp['id'])
                    if not emp_days:
                        continue
                    row = [emp['full_name'], emp['position'] or ""]
                    total_hours = 0.0
                    days_worked = 0
                    for day in range(1, days_in_month + 1):
                        entry = emp_days.get(day)
                        if not entry:
                            row.append("")
                            continue
                        hours_val = entry.get('hours', 0)
                        row.append(f"{hours_val:.1f}")
                        total_hours += hours_val
                        days_worked += 1
                    row.append(total_hours)
                    row.append(days_worked)
                    ws.append(row)

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file

    async def create_excel_report(self):
        """Создание Excel отчета с разбивкой по объектам и общим листом"""
        return await self._build_workbook(include_objects=True)

    async def create_overall_report(self):
        """Создание общего табеля без разбивки по объектам"""
        return await self._build_workbook(include_objects=False)

    async def build_action_log_report(self, start_date: date, end_date: date) -> Optional[io.BytesIO]:
        """Формирует выгрузку действий и геолокаций за период"""
        async with self.pool.acquire() as conn:
            time_logs = await conn.fetch(
                '''
                SELECT tl.date, tl.check_in, tl.check_out, tl.hours_worked, tl.status,
                       tl.check_in_lat, tl.check_in_lon, tl.check_out_lat, tl.check_out_lon,
                       e.full_name, e.position, e.is_admin, o.name as object_name
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                LEFT JOIN objects o ON tl.object_id = o.id
                WHERE tl.date BETWEEN $1 AND $2
                ORDER BY tl.date DESC, e.full_name
                ''',
                start_date,
                end_date,
            )
            location_attempts = await conn.fetch(
                '''
                SELECT la.message_time, la.action, la.latitude, la.longitude, la.distance, la.accuracy,
                       la.is_suspicious, la.reason,
                       e.full_name, e.position, e.is_admin,
                       o.name AS object_name
                FROM location_audit la
                LEFT JOIN employees e ON la.employee_id = e.id
                LEFT JOIN objects o ON la.object_id = o.id
                WHERE la.message_time::date BETWEEN $1 AND $2
                ORDER BY la.message_time DESC
                ''',
                start_date,
                end_date,
            )

        if not time_logs and not location_attempts:
            return None

        wb = Workbook()
        ws_actions = wb.active
        ws_actions.title = "Отметки"
        ws_actions.append(
            [
                "Дата",
                "Сотрудник",
                "Должность",
                "Роль",
                "Объект",
                "Статус",
                "Приход",
                "Широта прихода",
                "Долгота прихода",
                "Уход",
                "Широта ухода",
                "Долгота ухода",
                "Часы",
            ]
        )

        for log in time_logs:
            role = "Админ" if log['is_admin'] else "Сотрудник"
            ws_actions.append(
                [
                    log['date'].strftime("%Y-%m-%d"),
                    log['full_name'],
                    log['position'] or "",
                    role,
                    log['object_name'] or "",
                    log['status'] or "",
                    log['check_in'].strftime("%H:%M:%S") if log['check_in'] else "",
                    float(log['check_in_lat']) if log['check_in_lat'] is not None else "",
                    float(log['check_in_lon']) if log['check_in_lon'] is not None else "",
                    log['check_out'].strftime("%H:%M:%S") if log['check_out'] else "",
                    float(log['check_out_lat']) if log['check_out_lat'] is not None else "",
                    float(log['check_out_lon']) if log['check_out_lon'] is not None else "",
                    float(log['hours_worked'] or 0),
                ]
            )

        ws_geo = wb.create_sheet("Геолокации")
        ws_geo.append(
            [
                "Время (МСК)",
                "Сотрудник",
                "Должность",
                "Роль",
                "Действие",
                "Объект",
                "Широта",
                "Долгота",
                "Расстояние до объекта, м",
                "Точность, м",
                "Пометка",
                "Причина",
            ]
        )
        for attempt in location_attempts:
            role = "Админ" if attempt['is_admin'] else "Сотрудник"
            ws_geo.append(
                [
                    attempt['message_time'].strftime("%Y-%m-%d %H:%M:%S") if attempt['message_time'] else "",
                    attempt['full_name'] or "",
                    attempt['position'] or "",
                    role,
                    attempt['action'] or "",
                    attempt['object_name'] or "",
                    float(attempt['latitude']) if attempt['latitude'] is not None else "",
                    float(attempt['longitude']) if attempt['longitude'] is not None else "",
                    attempt['distance'] if attempt['distance'] is not None else "",
                    float(attempt['accuracy']) if attempt['accuracy'] is not None else "",
                    "Сомнительно" if attempt['is_suspicious'] else "Ок",
                    attempt['reason'] or "",
                ]
            )

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file
    
    async def send_timesheet_email(self, callback: types.CallbackQuery):
        """Отправка общего табеля в чат администратора"""
        excel_file = await self.create_overall_report()
        if not excel_file:
            await callback.message.answer("Не удалось сформировать табель")
            return

        await self.bot.send_document(
            chat_id=callback.from_user.id,
            document=types.BufferedInputFile(
                excel_file.getvalue(),
                filename=f"Общий_табель_{now_moscow().strftime('%Y_%m_%d')}.xlsx",
            ),
            caption="Общий табель без разбивки по объектам",
        )

    async def handle_sick_clear(self, message: types.Message, state: FSMContext):
        """Снятие больничного на сегодня"""
        user_id = message.from_user.id
        if not await self.check_access(user_id):
            await message.answer("Доступ запрещен.")
            return

        today = self.moscow_today()
        async with self.pool.acquire() as conn:
            sick_today = await conn.fetchval(
                '''
                SELECT id FROM time_logs
                WHERE employee_id = (SELECT id FROM employees WHERE telegram_id = $1)
                  AND date = $2 AND status = 'sick'
                LIMIT 1
                ''',
                user_id,
                today,
            )

            if not sick_today:
                await message.answer(
                    "На сегодня больничный не установлен.",
                    reply_markup=await self.build_main_keyboard_for_user(user_id),
                )
                return

            await conn.execute(
                '''
                DELETE FROM time_logs
                WHERE id = $1
                ''',
                sick_today,
            )

        await message.answer(
            "Больничный снят. Можно отмечать приход.",
            reply_markup=await self.build_main_keyboard_for_user(user_id),
        )
        await state.clear()

    async def send_timesheet_to_emails(self, subject: Optional[str] = None) -> bool:
        """Заглушка отправки по email (отложено)"""
        logger.info("Отправка по email пока отключена")
        return False

    def _build_email_message(self, excel_file: io.BytesIO, subject: Optional[str]) -> MIMEMultipart:
        message = MIMEMultipart()
        message['From'] = Config.SMTP_USERNAME
        message['To'] = ', '.join(Config.EMAIL_RECIPIENTS)
        message['Subject'] = subject or f"Табель {now_moscow().strftime('%B %Y')}"

        message.attach(MIMEText("Табель учета рабочего времени во вложении", 'plain'))

        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(excel_file.getvalue())
        encoders.encode_base64(attachment)
        attachment.add_header(
            'Content-Disposition',
            f'attachment; filename="Табель_{now_moscow().strftime("%Y_%m")}.xlsx"'
        )
        message.attach(attachment)
        return message

    async def send_action_log(self, callback: types.CallbackQuery):
        """Отправка журнала действий и геолокаций за текущий месяц"""
        today = self.moscow_today()
        start_date = date(today.year, today.month, 1)
        excel_file = await self.build_action_log_report(start_date, today)
        if not excel_file:
            await callback.message.answer("Нет данных для выгрузки за выбранный период.")
            return

        filename = f"Логи_{start_date.strftime('%Y_%m_%d')}_{today.strftime('%Y_%m_%d')}.xlsx"
        await self.bot.send_document(
            chat_id=callback.from_user.id,
            document=types.BufferedInputFile(
                excel_file.getvalue(),
                filename=filename,
            ),
            caption="Отчет по действиям и геолокациям (текущий месяц)",
        )

    async def _send_email_async(self, message: MIMEMultipart):
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, self._send_email_sync, message)

    def _send_email_sync(self, message: MIMEMultipart):
        with smtplib.SMTP(Config.SMTP_SERVER, Config.SMTP_PORT) as server:
            server.starttls()
            server.login(Config.SMTP_USERNAME, Config.SMTP_PASSWORD)
            server.send_message(message)
    
    async def show_employees(self, callback: types.CallbackQuery):
        """Показать список сотрудников"""
        async with self.pool.acquire() as conn:
            # Гарантируем наличие столбца position даже на старых БД
            await self.ensure_column(conn, "employees", "position", "VARCHAR(100)")

            employees = await conn.fetch('''
                SELECT full_name, position, telegram_id, is_admin,
                       (SELECT COUNT(*) FROM time_logs WHERE employee_id = employees.id
                        AND EXTRACT(MONTH FROM date) = EXTRACT(MONTH FROM CURRENT_DATE)) as days_worked
                FROM employees WHERE is_approved = TRUE ORDER BY full_name
            ''')
        
        text = "Сотрудники:\n\n"
        for emp in employees:
            role = "Админ" if emp['is_admin'] else "Сотрудник"
            position = emp['position'] or "Не указана"
            text += f"{emp['full_name']}\n"
            text += f"Должность: {position}\n"
            text += f"Роль: {role}, ID: {emp['telegram_id']}, Дней: {emp['days_worked']}\n\n"

        await callback.message.answer(text)

    async def show_today_checkins(self, callback: types.CallbackQuery):
        """Отобразить сегодняшние приходы"""
        today = self.moscow_today()
        async with self.pool.acquire() as conn:
            rows = await conn.fetch(
                '''
                SELECT e.full_name, tl.check_in, o.name as object_name
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                LEFT JOIN objects o ON tl.object_id = o.id
                WHERE tl.date = $1 AND tl.check_in IS NOT NULL
                ORDER BY tl.check_in
                ''',
                today,
            )

        if not rows:
            await callback.message.answer("Сегодня еще никто не отметился на приход.")
            return

        lines = ["Приходы сегодня:"]
        for row in rows:
            object_part = f" ({row['object_name']})" if row['object_name'] else ""
            lines.append(f"{row['full_name']} — {row['check_in'].strftime('%H:%M')}{object_part}")

        async with self.pool.acquire() as conn:
            missing_checkouts = await conn.fetch(
                '''
                SELECT e.full_name
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                WHERE tl.date = $1 AND tl.check_in IS NOT NULL AND tl.check_out IS NULL
                  AND e.is_active = TRUE AND e.is_approved = TRUE
                ORDER BY e.full_name
                ''',
                today,
            )

        if missing_checkouts:
            lines.append("\nНе отметили уход:")
            for row in missing_checkouts:
                lines.append(f"• {row['full_name']}")

        await callback.message.answer("\n".join(lines))
    
    async def show_objects_admin(self, callback: types.CallbackQuery):
        """Показать объекты"""
        async with self.pool.acquire() as conn:
            objects = await conn.fetch('SELECT name, address, latitude, longitude FROM objects ORDER BY name')
        
        text = "Объекты:\n\n"
        for obj in objects:
            text += f"{obj['name']}\n"
            if obj['address']:
                text += f"Адрес: {obj['address']}\n"
            if obj['latitude']:
                text += f"Координаты: {obj['latitude']:.4f}, {obj['longitude']:.4f}\n"
            text += "\n"
        
        await callback.message.answer(text)
    
    async def show_stats(self, callback: types.CallbackQuery):
        """Показать статистику"""
        today = self.moscow_today()
        month_start = date(today.year, today.month, 1)
        
        async with self.pool.acquire() as conn:
            employees = await conn.fetch(
                '''
                SELECT id, full_name
                FROM employees
                WHERE is_active = TRUE AND is_approved = TRUE
                ORDER BY full_name
                '''
            )

            month_rows = await conn.fetch(
                '''
                SELECT tl.employee_id,
                       SUM(tl.hours_worked) AS hours,
                       COUNT(DISTINCT tl.date) FILTER (WHERE tl.status <> 'sick') AS days_worked,
                       COUNT(*) FILTER (WHERE tl.status = 'sick') AS sick_days
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                WHERE tl.date >= $1 AND tl.date <= $2
                  AND e.is_active = TRUE AND e.is_approved = TRUE
                GROUP BY tl.employee_id
                ''',
                month_start,
                today,
            )

            today_stats = await conn.fetchrow('''
                SELECT 
                    COUNT(DISTINCT employee_id) as worked_today,
                    SUM(hours_worked) as hours_today
                FROM time_logs WHERE date = $1
            ''', today)
            missing_checkins = await conn.fetch(
                '''
                SELECT e.full_name
                FROM employees e
                WHERE e.is_active = TRUE AND e.is_approved = TRUE
                  AND NOT EXISTS (
                      SELECT 1 FROM time_logs tl
                      WHERE tl.employee_id = e.id AND tl.date = $1 AND tl.check_in IS NOT NULL
                  )
                  AND NOT EXISTS (
                      SELECT 1 FROM time_logs tl
                      WHERE tl.employee_id = e.id AND tl.date = $1 AND tl.status = 'sick'
                  )
                ORDER BY e.full_name
                ''',
                today,
            )
            missing_checkouts = await conn.fetch(
                '''
                SELECT e.full_name
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                WHERE tl.date = $1 AND tl.check_in IS NOT NULL AND tl.check_out IS NULL
                  AND e.is_active = TRUE AND e.is_approved = TRUE
                ORDER BY e.full_name
                ''',
                today,
            )
        
        text = f"Статистика за {today.strftime('%B %Y')}:\n\n"
        text += f"Активных сотрудников: {total_employees}\n"
        text += f"Отметились в месяце: {worked_count} ({worked_pct:.0f}%)\n"
        text += f"Всего часов: {total_hours:.1f}\n"
        text += f"Рабочих дней (суммарно): {total_work_days}\n"
        text += f"Больничных дней: {total_sick_days}\n"

        text += "\nТоп по часам:\n"
        if not top_hours:
            text += "—\n"
        else:
            for row in top_hours:
                text += f"• {next(emp['full_name'] for emp in employees if emp['id'] == row['employee_id'])} — {float(row['hours'] or 0):.1f} ч, дней: {int(row['days_worked'] or 0)}\n"

        text += "\nБольничные за месяц:\n"
        if not sick_leaders:
            text += "—\n"
        else:
            for row in sick_leaders:
                text += f"• {next(emp['full_name'] for emp in employees if emp['id'] == row['employee_id'])} — {int(row['sick_days'] or 0)} дн.\n"

        if non_working:
            text += "\nБез отметок в этом месяце:\n"
            for name in non_working:
                text += f"• {name}\n"
        else:
            text += "\nВсе сотрудники отмечались в этом месяце.\n"

        text += "\nСегодня:\n"
        text += f"Работали: {today_stats['worked_today'] or 0}\n"
        text += f"Часов: {today_stats['hours_today'] or 0:.1f}\n"
        text += "Не отметили приход:\n"
        if not missing_checkins:
            text += "—\n"
        else:
            for row in missing_checkins:
                text += f"• {row['full_name']}\n"
        text += "Не отметили уход:\n"
        if not missing_checkouts:
            text += "—"
        else:
            for row in missing_checkouts:
                text += f"• {row['full_name']}\n"
        
        await callback.message.answer(text)
    
    # Планировщик
    async def setup_scheduler(self):
        """Настройка планировщика задач"""
        self.scheduler.add_job(
            self.remind_checkin,
            CronTrigger(hour=8, minute=50),
            id="remind_checkin_1"
        )
        self.scheduler.add_job(
            self.remind_checkout,
            CronTrigger(hour=18, minute=0),
            id="reminder"
        )
        self.scheduler.add_job(
            self.auto_checkout,
            CronTrigger(hour=20, minute=0),
            id="auto_checkout"
        )
        self.scheduler.add_job(
            self.force_checkout_end_of_day,
            CronTrigger(hour=23, minute=59),
            id="force_checkout_end_of_day"
        )
        self.scheduler.add_job(
            self.send_periodic_timesheet,
            CronTrigger(hour=20, minute=0),
            id="periodic_timesheet"
        )
        self.scheduler.start()
        logger.info("Планировщик запущен")

    async def remind_checkin(self):
        """Напоминание о приходе для сотрудников"""
        today = self.moscow_today()
        if not self.is_working_day(today):
            logger.info("Пропускаем напоминание о приходе в выходной/праздничный день: %s", today)
            return
        async with self.pool.acquire() as conn:
            employees = await conn.fetch(
                '''
                SELECT e.telegram_id, e.full_name, e.is_admin
                FROM employees e
                WHERE e.is_active = TRUE AND e.is_approved = TRUE
                  AND NOT EXISTS (
                      SELECT 1 FROM time_logs tl
                      WHERE tl.employee_id = e.id AND tl.date = $1 AND tl.check_in IS NOT NULL
                  )
                  AND NOT EXISTS (
                      SELECT 1 FROM time_logs tl
                      WHERE tl.employee_id = e.id AND tl.date = $1 AND tl.status = 'sick'
                  )
                ''' ,
                today,
            )

        for emp in employees:
            try:
                await self.bot.send_message(
                    emp['telegram_id'],
                    "Не забудьте отметить приход!",
                    reply_markup=self.get_main_keyboard(is_admin=emp['is_admin']),
                )
            except Exception as e:
                logger.error(f"Ошибка отправки напоминания о приходе: {e}")
    
    async def remind_checkout(self):
        """Напоминание об уходе"""
        today = self.moscow_today()
        if not self.is_working_day(today):
            logger.info("Пропускаем напоминание об уходе в выходной/праздничный день: %s", today)
            return
        
        async with self.pool.acquire() as conn:
            employees = await conn.fetch('''
                SELECT e.telegram_id, e.full_name, e.is_admin
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                WHERE tl.date = $1 AND tl.check_in IS NOT NULL AND tl.check_out IS NULL
                AND e.is_active = TRUE
            ''', today)
        
        for emp in employees:
            try:
                await self.bot.send_message(
                    emp['telegram_id'],
                    "Напоминание! Не забудьте отметить уход.",
                    reply_markup=self.get_main_keyboard(is_admin=emp['is_admin'])
                )
            except Exception as e:
                logger.error(f"Ошибка отправки напоминания: {e}")
    
    async def auto_checkout(self):
        """Автоматический уход в 20:00 с ограничением 8 часов"""
        today = self.moscow_today()
        if not self.is_working_day(today):
            logger.info("Пропускаем авто-уход в выходной/праздничный день: %s", today)
            return

        async with self.pool.acquire() as conn:
            logs = await conn.fetch('''
                SELECT tl.id, e.telegram_id, e.full_name, e.is_admin, tl.check_in
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                WHERE tl.date = $1 AND tl.check_in IS NOT NULL AND tl.check_out IS NULL
            ''', today)
        
        now = self.moscow_now_naive()
        for log in logs:
            try:
                hours = (now - log['check_in']).seconds / 3600
                hours = min(max(0, math.ceil(hours)), Config.MAX_WORK_HOURS)

                async with self.pool.acquire() as conn:
                    await conn.execute('''
                        UPDATE time_logs SET check_out = $1, hours_worked = $2 WHERE id = $3
                    ''', now, hours, log['id'])
                
                await self.bot.send_message(
                    log['telegram_id'],
                    f"Автоматический уход в {now.strftime('%H:%M')}\nОтработано: {hours} ч.",
                    reply_markup=self.get_main_keyboard(is_admin=log['is_admin'])
                )
                logger.info(
                    "Авто-уход: %s (%s) часов %.0f в %s",
                    log['telegram_id'],
                    log['id'],
                    hours,
                    now.strftime('%H:%M'),
                )
            except Exception as e:
                logger.error(f"Ошибка автоматического ухода: {e}")

    async def force_checkout_end_of_day(self):
        """Автоматический уход в 23:59, если сотрудник не отметился (ограничение 8 часов)"""
        today = self.moscow_today()

        async with self.pool.acquire() as conn:
            logs = await conn.fetch(
                '''
                SELECT tl.id, e.telegram_id, e.full_name, e.is_admin, tl.check_in
                FROM time_logs tl
                JOIN employees e ON tl.employee_id = e.id
                WHERE tl.date = $1 AND tl.check_in IS NOT NULL AND tl.check_out IS NULL
                ''',
                today,
            )

        if not logs:
            return

        cutoff = datetime.combine(today, time(23, 59)).replace(tzinfo=None)

        for log in logs:
            try:
                hours = 0
                if log['check_in']:
                    hours = (cutoff - log['check_in']).total_seconds() / 3600
                hours = min(max(0, math.ceil(hours)), Config.MAX_WORK_HOURS)

                async with self.pool.acquire() as conn:
                    await conn.execute(
                        '''
                        UPDATE time_logs
                        SET check_out = $1, hours_worked = $2
                        WHERE id = $3
                        ''',
                        cutoff,
                        hours,
                        log['id'],
                    )

                await self.bot.send_message(
                    log['telegram_id'],
                    f"День закрыт автоматически в 23:59\nОтработано: {hours} ч.",
                    reply_markup=self.get_main_keyboard(is_admin=log['is_admin']),
                )
                logger.info(
                    "Ночной авто-уход: %s (%s) часов %.0f в 23:59",
                    log['telegram_id'],
                    log['id'],
                    hours,
                )
            except Exception as e:
                logger.error(f"Ошибка ночного авто-ухода: {e}")

    async def send_periodic_timesheet(self):
        """Автоматическая отправка табеля дважды в месяц"""
        today = self.moscow_today()
        last_day = monthrange(today.year, today.month)[1]

        if today.day not in (15, last_day):
            return

        try:
            excel_file = await self.create_overall_report()
            if not excel_file:
                logger.warning("Не удалось сформировать табель для плановой отправки")
                return

            for admin_id in Config.ADMIN_IDS:
                try:
                    await self.bot.send_document(
                        chat_id=admin_id,
                        document=types.BufferedInputFile(
                            excel_file.getvalue(),
                            filename=f"Общий_табель_{today.strftime('%Y_%m_%d')}.xlsx",
                        ),
                        caption="Плановая отправка общего табеля",
                    )
                except Exception as send_err:
                    logger.error(f"Не удалось отправить табель администратору {admin_id}: {send_err}")
        except Exception as e:
            logger.error(f"Ошибка плановой отправки табеля: {e}")
    
    # Запуск
    async def run(self):
        """Запуск бота"""
        await self.init_db()
        await self.setup_scheduler()
        logger.info("Бот запущен")
        await self.dp.start_polling(self.bot)

# Точка входа
if __name__ == "__main__":
    bot = WorkTimeBot()
    
    try:
        asyncio.run(bot.run())
    except KeyboardInterrupt:
        logger.info("Бот остановлен")
